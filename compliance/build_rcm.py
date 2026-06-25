from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
DATE = "2026-06-22"

# ---- palette ----
NAVY = "1F4E78"; STEEL = "2E5C8A"; LIGHT = "DCE6F1"; BAND = "F2F6FB"
GREEN_F = "C6EFCE"; GREEN_T = "006100"
AMBER_F = "FFEB9C"; AMBER_T = "9C6500"
RED_F = "FFC7CE"; RED_T = "9C0006"
GREY = "808080"

def f(sz=10, bold=False, color="000000"): return Font(name=FONT, size=sz, bold=bold, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAPC = Alignment(wrap_text=True, vertical="center", horizontal="center")

wb = Workbook()

# ============================================================== COVER
cov = wb.active; cov.title = "Cover"
cov.sheet_view.showGridLines = False
cov.column_dimensions["A"].width = 3
cov.column_dimensions["B"].width = 30
cov.column_dimensions["C"].width = 95
def crow(r, label, val, lbold=True):
    cov[f"B{r}"] = label; cov[f"B{r}"].font = f(10, lbold, NAVY); cov[f"B{r}"].alignment = WRAP
    cov[f"C{r}"] = val; cov[f"C{r}"].font = f(10); cov[f"C{r}"].alignment = WRAP
cov["B2"] = "Risk & Control Matrix (RCM)"; cov["B2"].font = f(20, True, NAVY)
cov["B3"] = "SOX / COSO 2013 Internal-Control Readiness — IT General Controls & Application Controls"; cov["B3"].font = f(11, False, STEEL)
for r in (2,3): cov.merge_cells(f"B{r}:C{r}")
crow(6, "Entity", "Invisible Consulting — Oshinei Enterprise ERP")
crow(7, "System in scope", "Invisible ERP V2 (NestJS API · Next.js web · Drizzle · PostgreSQL, multi-tenant w/ RLS)")
crow(8, "Filer status", "Emerging Growth Company (EGC) — JOBS Act")
crow(9, "Framework", "COSO 2013 (5 components / 17 principles) · ITGC aligned to COBIT · PCAOB AS 2201")
crow(10, "Regulatory driver", "SOX 302 (CEO/CFO certification) · SOX 404(a) (management ICFR assessment)")
crow(11, "Prepared (draft)", DATE + "  ·  Version 1.0 (DRAFT for review with external auditor + SOX advisor)")
crow(12, "Owner", "SOX PMO / Controller (to assign)")

cov["B15"] = "EGC implications (read before scoping)"; cov["B15"].font = f(11, True, NAVY)
egc = ("• As an EGC, SOX 404(b) AUDITOR ATTESTATION on ICFR is DEFERRED — up to 5 fiscal years post-IPO, or until EGC status is lost "
       "(>$1.235B revenue, large-accelerated-filer status, >$1B non-convertible debt/3yr).\n"
       "• STILL REQUIRED from the EGC's first/second annual report: SOX 302 CEO/CFO certifications and SOX 404(a) MANAGEMENT's assessment of ICFR.\n"
       "• Your INDEPENDENT FINANCIAL AUDITOR still tests IT-dependent and key controls to the extent they rely on them in the financial-statement audit — "
       "so this RCM and the underlying ITGC must be real, documented, and operating, attestation deferral notwithstanding.\n"
       "• Practical effect: build now; the deferral lowers the formal-attestation bar, NOT the bar for sound, evidenced internal control.")
cov["B16"] = egc; cov.merge_cells("B16:C20"); cov["B16"].font = f(10); cov["B16"].alignment = WRAP

cov["B22"] = "Status legend"; cov["B22"].font = f(11, True, NAVY)
leg = [("Implemented", GREEN_F, GREEN_T, "Control exists in the system/process today (file reference given). Needs documentation + testing of operating effectiveness."),
       ("Partial", AMBER_F, AMBER_T, "Control partially present — capability or component exists but must be formalized, enforced, or extended."),
       ("Gap", RED_F, RED_T, "Control not yet implemented — on the remediation plan (see 'Gap Remediation' tab).")]
for i,(lab,fc,tc,desc) in enumerate(leg):
    r = 23+i
    cov[f"B{r}"] = lab; cov[f"B{r}"].fill = fill(fc); cov[f"B{r}"].font = f(10, True, tc); cov[f"B{r}"].alignment = Alignment(horizontal="center")
    cov[f"C{r}"] = desc; cov[f"C{r}"].font = f(10); cov[f"C{r}"].alignment = WRAP
cov["B27"] = ("Disclaimer: This RCM is an engineering/IT-control working draft to accelerate audit readiness. It is not legal, accounting, or "
              "SOX-compliance advice. Validate scope, control selection, and testing with your independent audit firm and a qualified SOX advisor.")
cov.merge_cells("B27:C29"); cov["B27"].font = f(9, False, GREY); cov["B27"].alignment = WRAP

# ============================================================== RCM
HEAD = ["Control ID","Cycle / Process","Category","FSLI / Significant Account","Risk — What Could Go Wrong",
        "Assertion(s)","Control Description","Prev / Det","Nature","Frequency","Control Owner","COSO Principle",
        "System / Code Reference","Test of Design (TOD)","Test of Operating Effectiveness (TOE)","Key Evidence","Status"]
W = [11,17,12,20,30,13,46,9,14,13,16,12,30,34,34,26,12]

# Each row: the 17 columns above.
R = []
def add(*cols): R.append(list(cols))

# ---- Entity-Level Controls ----
add("ELC-01","Entity-Level","Entity","All","Tone-at-the-top / ethics breaches enable override of controls and fraud.","—",
    "Code of conduct & ethics policy issued, acknowledged annually by all staff; conflict-of-interest declarations.","Prev","Manual","Annual","CEO / HR","P1",
    "Policy (non-system)","Inspect signed policy + acknowledgement register.","Re-perform: sample staff have current signed acknowledgement.","Signed policy register","Gap")
add("ELC-02","Entity-Level","Entity","All","No independent oversight of financial reporting / ICFR.","—",
    "Audit committee charter; quarterly review of financial reporting, ICFR status and audit findings.","Det","Manual","Quarterly","Board / Audit Cttee","P2",
    "Governance (non-system)","Inspect charter + minutes.","Inspect minutes evidencing ICFR oversight each quarter.","AC minutes","Gap")
add("ELC-03","Entity-Level","Entity","All","Transactions executed without proper authority.","Authorization",
    "Documented delegation-of-authority / approval matrix mapped to system roles & limits.","Prev","Manual","Annual","CFO","P3",
    "DoA matrix ↔ permissions.ts","Inspect DoA matrix; reconcile to RBAC roles.","Sample approvals respect DoA thresholds.","Approval matrix","Gap")
add("ELC-04","Entity-Level","Entity","All","Misconduct undetected; no safe reporting channel.","—",
    "Anonymous whistleblower / ethics hotline with non-retaliation policy and case tracking.","Det","Manual","Continuous","Audit Cttee","P1",
    "External hotline","Inspect hotline setup + policy.","Inspect case log + resolution evidence.","Hotline log","Gap")
add("ELC-05","Entity-Level","Entity","All","Fraud risks not identified / mitigated.","—",
    "Annual fraud risk assessment covering revenue, cash, JE override, related parties.","Det","Manual","Annual","CFO / IA","P8",
    "Fraud risk register","Inspect risk assessment.","Inspect that identified risks map to controls in this RCM.","Fraud matrix","Gap")
add("ELC-06","Entity-Level","Entity","All","Errors in results undetected by management.","Accuracy",
    "Management review of actuals vs budget/forecast with documented investigation of variances.","Det","Manual","Monthly","CFO / Controller","P16",
    "budget, bi modules","Inspect budget vs actual reports.","Sample months: evidence of review + variance follow-up.","Variance review","Partial")

# ---- ITGC: Access to Programs & Data ----
add("ITGC-AC-01","ITGC · Access","ITGC","All","Unauthenticated access to financial data/functions.","Restricted access",
    "Global JWT auth guard — every endpoint requires a valid token/API key unless explicitly @Public.","Prev","Automated","Continuous","Eng Lead","P11",
    "common/guards.ts (JwtAuthGuard); app.module.ts","Inspect global guard registration; attempt unauthenticated call → 401.","Re-perform on endpoint sample; review access logs over period.","Config + 401 test","Implemented")
add("ITGC-AC-02","ITGC · Access","ITGC","All","Users perform actions beyond their role; destructive ops over-granted.","Authorization",
    "RBAC — 37 fine-grained permissions; @Permissions enforced; refund/void staff-only; fiscal close = exec.","Prev","Automated","Continuous","Eng Lead / Controller","P10/P11",
    "guards.ts (PermissionsGuard); packages/shared/permissions.ts; payments.controller.ts; ledger.controller.ts","Inspect role→perm map; attempt over-privileged call → 403.","Re-perform across role matrix; tie to UAR (AC-08).","Perm map + 403 tests","Implemented")
add("ITGC-AC-03","ITGC · Access","ITGC","All","Cross-tenant data leakage / mis-posting.","Restricted access",
    "PostgreSQL Row-Level Security; SET LOCAL ROLE app_user + tenant pin; USING+WITH CHECK; fail-closed in prod.","Prev","Automated","Continuous","Eng Lead","P11",
    "common/tenant-tx.interceptor.ts; drizzle/0002_rls.sql; 0003_tenants_rls.sql","Inspect RLS policies; attempt cross-tenant read/insert → blocked.","Re-perform cross-tenant tests; review for RLS-disabled paths.","RLS policy + test","Implemented")
add("ITGC-AC-04","ITGC · Access","ITGC","All / PII","Secrets/PII exposed at rest.","Confidentiality",
    "AES-256-GCM encryption-at-rest for secrets (TOTP seeds, webhook secrets); APP_ENC_KEY required in prod.","Prev","Automated","Continuous","Eng Lead","P11",
    "common/crypto.ts","Inspect cipher + fail-closed key gate.","Confirm prod boot fails without APP_ENC_KEY; sample encrypted rows.","Code + boot test","Implemented")
add("ITGC-AC-05","ITGC · Access","ITGC","All","Machine credential escalates to admin / HQ bypass.","Authorization",
    "API keys are downscoped (never Admin/HQ-bypass) and RLS-bound to their own tenant.","Prev","Automated","Continuous","Eng Lead","P11",
    "guards.ts (api-key path); platform/api-key.service.ts","Inspect key→role mapping (role='Sales', no bypass).","Attempt admin action with API key → denied.","Code + test","Implemented")
add("ITGC-AC-06","ITGC · Access","ITGC","All","Stolen password yields full access to finance functions.","Restricted access",
    "Multi-factor authentication (TOTP) enforced for Admin and finance roles.","Prev","Automated","Continuous","IT Security","P11",
    "auth.service.ts (login TOTP gate; setup/enable/disable); permissions.ts (requiresMfa policy); crypto.ts (AES-256-GCM seed); cutover/compliance.ts (ToE)","Inspect MFA enforcement policy for privileged roles.","Sample privileged logins require 2nd factor (MFA-enabled login w/o code → 401; un-enrolled privileged user flagged must_setup_mfa) — re-performed by the harness.","MFA config","Implemented")
add("ITGC-AC-07","ITGC · Access","ITGC","All","Session token stealable via XSS; weak/stale credentials; sessions never expire.","Restricted access",
    "Web session is an httpOnly JWT cookie (token unreadable by JS → XSS can't exfiltrate it) + CSRF double-submit on cookie-authenticated mutations + a web Content-Security-Policy; plus password policy, forced first-login change, and JWT/session expiry. Bearer/API-key clients are CSRF-exempt (no ambient cookie).","Prev","Automated","Continuous","IT Security","P11",
    "common/cookies.ts; common/guards.ts (cookie auth + CSRF); auth.controller.ts (login/logout set/clear cookies); apps/web/next.config.mjs (CSP); 0045_must_change_password; cutover/cookie-auth.ts (ToE)","Inspect cookie flags (HttpOnly/SameSite), CSRF enforcement, and CSP.","Sample: cookie auth works; cookie mutation w/o X-CSRF-Token → 403; Bearer exempt; logout clears; expired token rejected (re-performed by the harness).","Auth config; cookie-auth ToE","Implemented")
add("ITGC-AC-08","ITGC · Access","ITGC","All","Access creep; terminated users retain access.","Restricted access",
    "Quarterly User Access Review — recertify every user × permission; remove on termination.","Det","Manual","Quarterly","Controller / IT","P11/P16",
    "admin-users.service.ts (access-review / export CSV / certify); admin-users.controller.ts; cutover/compliance.ts (ToE)","Inspect UAR procedure + sample sign-off.","Re-perform: each quarter reviewed & exceptions remediated (automated harness re-performs report/export/certify).","UAR sign-off","Implemented")
add("ITGC-AC-09","ITGC · Access","ITGC","All","Conflicting duties in one user enable fraud (e.g., post & approve JE).","Authorization",
    "Segregation-of-Duties conflict ruleset with preventive blocks + detective conflict report.","Prev/Det","Auto+Manual","Quarterly","Controller / IT","P10/P11",
    "permissions.ts (SOD_RULES, 13 rules); admin-users.service.ts (assertNoSodConflict preventive block); sod.service.ts (detective report); cutover/compliance.ts (ToE)","Inspect SoD rule set vs role design.","Run conflict report; assigning a conflicting set is blocked (422) unless justified-override+reason — re-performed by the harness.","SoD report","Implemented")
add("ITGC-AC-10","ITGC · Access","ITGC","All","No record of who changed financial data.","Completeness",
    "Audit trail — every mutating request logged with user, timestamp, IP, action, status (field-level before/after is ITGC-AC-14).","Det","Automated","Continuous","Eng Lead","P13/P16",
    "common/audit.interceptor.ts; status-log.service.ts; drizzle/0062_audit_log_immutable.sql (append-only trigger); cutover/compliance.ts (ToE)","Inspect logged actions/fields; confirm coverage of GL/AR/AP/cash; verify append-only trigger.","Sample transactions traced to audit log; UPDATE/DELETE on audit_log rejected (re-performed by the harness).","Audit log","Implemented")
add("ITGC-AC-11","ITGC · Access","ITGC","Revenue / Cash","Past financial records altered/deleted undetectably.","Integrity",
    "Append-only, per-tenant HASH-CHAINED electronic journal — altering any row breaks later hashes (tamper-evident).","Prev/Det","Automated","Continuous","Eng Lead","P13",
    "pos-fiscal/journal.service.ts (prevHash→hash chain, FOR UPDATE)","Inspect chaining logic; recompute a chain segment.","Re-verify chain integrity over a period; test tamper detection.","Chain verify","Implemented")
add("ITGC-AC-12","ITGC · Access","ITGC","All","Secrets hardcoded/shared/un-rotated → compromise.","Confidentiality",
    "Secrets (JWT_SECRET, APP_ENC_KEY, PSP_WEBHOOK_SECRET, DB creds) held in KMS/vault; rotated; no dev fallbacks in prod.","Prev","Manual","Continuous","DevOps / Security","P11",
    "env.validation.ts (fail-closed boot gate); docs/ops/secrets.md","Inspect secret storage + rotation policy.","Confirm no plaintext secrets in repo/CI; prod boot blocked without secrets; rotation evidence.","Vault config + env.validation","Implemented")
add("ITGC-AC-13","ITGC · Access","ITGC","All","Direct DB access bypasses app controls; shared DB superuser.","Restricted access",
    "Named DB users, least privilege, app uses non-superuser role; DBA access logged & restricted.","Prev","Manual","Continuous","DBA / DevOps","P11",
    "tools/ops/sql/prod-db-roles.sql (non-owner ierp_app login); app_user role","Inspect DB role grants + access list.","Sample DB access reviewed; app connects as non-owner least-priv role; no shared superuser in app path.","DB grants","Implemented")
add("ITGC-AC-14","ITGC · Access","ITGC","Revenue / Cash; Expenditure","Financial data changed without a field-level record of WHAT changed (old→new).","Completeness/Integrity",
    "Field-level change log — DB triggers capture OLD/NEW row images + changed columns + actor on every INSERT/UPDATE/DELETE of the core financial tables (GL header, AP/AR sub-ledgers, AP payments, tenders); append-only.","Det","Automated","Continuous","Eng Lead","P13/P16",
    "drizzle/0116_field_change_log.sql (log_data_change triggers; data_change_log append-only); tenant-tx.interceptor.ts (app.actor GUC); audit-viewer.service.ts (changes / GET /api/admin/audit/changes); cutover/compliance.ts (ToE)","Inspect trigger coverage + actor capture; verify append-only trigger.","Sample a financial update → change log holds old/new + changed columns + actor; UPDATE/DELETE on data_change_log rejected (re-performed by the harness).","Field change log","Implemented")

# ---- ITGC: Change Management ----
add("ITGC-CM-01","ITGC · Change","ITGC","All","Unauthorized/untested code reaches production.","—",
    "All code changes via peer-reviewed PR; branch protection on main; no self-merge; CI must pass before merge.","Prev","Auto+Manual","Per change","Head of Eng","P11",
    ".github/rulesets/main-branch-protection.json; .github/CODEOWNERS","Inspect branch protection + required-review settings.","Sample merges: independent + code-owner approval + green CI; no force-push.","PR approvals","Implemented")
add("ITGC-CM-02","ITGC · Change","ITGC","All","Ad-hoc prod schema change corrupts data.","—",
    "Schema changes only via reviewed Drizzle migrations; no direct prod DDL; migration journal maintained.","Prev","Auto+Manual","Per change","Head of Eng / DBA","P11",
    "apps/api/drizzle/*; meta/_journal.json; CODEOWNERS on drizzle/","Inspect migration process + journal.","Sample migrations: code-owner reviewed + applied via pipeline only.","Migration log","Implemented")
add("ITGC-CM-03","ITGC · Change","ITGC","All","Developer self-deploys unreviewed change to prod.","—",
    "Segregation: developer ≠ prod deployer; production deploy requires separate approval.","Prev","Manual","Per change","Head of Eng","P11",
    ".github/workflows/deploy.yml (production environment required reviewers)","Inspect deploy approval gate + role separation.","Sample deploys: approver ≠ author.","Deploy approvals","Implemented")
add("ITGC-CM-04","ITGC · Change","ITGC","All","Changes not traceable to an authorized request.","—",
    "Change traceability: ticket → PR → test → deploy linkage retained for every change.","Det","Manual","Per change","Head of Eng","P11",
    ".github/pull_request_template.md (required linked ticket)","Inspect linkage convention.","Sample changes traced ticket→PR→deploy.","Traceability","Implemented")
add("ITGC-CM-05","ITGC · Change","ITGC","All","Emergency fixes bypass controls permanently.","—",
    "Emergency-change procedure with retroactive review/approval within defined SLA.","Prev","Manual","As needed","Head of Eng","P11",
    "docs/ops/change-management.md (emergency-change procedure)","Inspect emergency procedure.","Sample emergency changes had expedited review + 1-day retro.","Emergency log","Implemented")

# ---- ITGC: SDLC / Development ----
add("ITGC-SD-01","ITGC · SDLC","ITGC","All","New functionality goes live without design/test sign-off.","—",
    "SDLC policy: requirements, design, test and UAT sign-off prior to go-live.","Prev","Manual","Per project","Head of Eng / Product","P11",
    "SDLC policy (to author)","Inspect SDLC policy + a project's artefacts.","Sample releases: UAT + go-live sign-off present.","SDLC artefacts","Gap")
add("ITGC-SD-02","ITGC · SDLC","ITGC","All","Migrated opening balances incomplete/inaccurate.","Completeness/Accuracy",
    "Cutover data-migration controls: source→target balance reconciliation + sign-off; opening balances idempotent.","Prev/Det","Auto+Manual","At cutover","Controller / Eng","P11/P13",
    "tools/etl; tools/cutover; ledger opening-balances (idempotent on batch_ref)","Inspect migration reconciliation + sign-off.","Re-perform migrated trial-balance tie-out.","Migration recon","Partial")
add("ITGC-SD-03","ITGC · SDLC","ITGC","All","Control logic regresses unnoticed.","—",
    "Automated test suite is control evidence; CI archives dated results; key-control assertions covered.","Det","Automated","Per change","Head of Eng","P11/P16",
    "apps/api/test/unit.test.ts; vitest","Inspect tests map to key controls.","Review archived CI runs over period; all green.","CI reports","Partial")

# ---- ITGC: Computer Operations ----
add("ITGC-OP-01","ITGC · Operations","ITGC","All","Data loss with no recovery.","—",
    "Automated DB backups + a TESTED quarterly restore (recovery proven, not just scheduled).","Prev","Auto+Manual","Daily / Qtrly test","DevOps","P11",
    "tools/ops/pg-backup.sh, restore.sh, verify-restore.sh; BACKUP-RUNBOOK.md","Inspect backup schedule + scripted restore-drill procedure.","Inspect a successful restore-drill record each quarter.","Restore test","Implemented")
add("ITGC-OP-02","ITGC · Operations","ITGC","All","Prolonged outage; no continuity plan.","—",
    "DR / BCP plan with defined RTO/RPO and periodic test.","Prev","Manual","Annual test","CTO / DevOps","P11",
    "RTO/RPO in tools/ops/BACKUP-RUNBOOK.md; full DR/BCP to author","Inspect DR plan + RTO/RPO.","Inspect DR test results.","DR test","Partial")
add("ITGC-OP-03","ITGC · Operations","ITGC","All","Failures/incidents undetected.","—",
    "Monitoring + alerting (APM/errors) with on-call and an incident-management log.","Det","Automated","Continuous","DevOps","P11/P16",
    "observability/instrumentation.ts (OTel); Sentry; docs/ops/observability-incident.md; /healthz+/readyz probes","Inspect alerting config + incident process.","Sample alerts → incident tickets resolved.","Incident log","Implemented")
add("ITGC-OP-04","ITGC · Operations","ITGC","Revenue / GL","Scheduled financial jobs fail silently (missed billing/postings).","Completeness",
    "Batch-job monitoring (recurring billing, FX revaluation, subscriptions) with failure alerting + review.","Det","Automated","Per run","DevOps / Controller","P11",
    "billing; fx; service modules; alert spec in docs/ops/observability-incident.md","Inspect job inventory + alerting.","Sample runs: success logged; failures investigated.","Job logs","Partial")

# ---- Revenue & Cash ----
add("REV-01","Revenue & Cash","Application","Revenue; Cash","Invalid/garbage data posted (negatives, wrong types).","Accuracy/Validity",
    "Zod schema validation on all sales/tender inputs (qty>0, price>=0, amount>0); standard error envelope.","Prev","Automated","Per txn","Eng Lead","P10",
    "pos.controller.ts; payments.controller.ts; common/zod-validation.pipe.ts","Inspect schemas; submit invalid payload → 400.","Re-perform invalid-input sample → rejected.","Validation tests","Implemented")
add("REV-02","Revenue & Cash","Application","Cash","Retried/double-submitted tender charges customer twice.","Occurrence/Completeness",
    "Payment idempotency — same idempotency_key returns the original tender; unique index backstops the race.","Prev","Automated","Per txn","Eng Lead","P10",
    "payments.service.ts (recordTender); schema/payments.ts (ux_payments_idem); migration 0057","Inspect key handling + unique index.","Re-submit same key → single charge; concurrent test.","Idempotency test","Implemented")
add("REV-03","Revenue & Cash","Application","Cash","Funds captured at gateway but no record (orphaned charge).","Completeness",
    "Row persisted Pending BEFORE gateway capture; flipped Failed on error — no unrecorded captured funds.","Prev","Automated","Per txn","Eng Lead","P10",
    "payments.service.ts (recordTender)","Inspect pre-persist + error path.","Simulate capture/persist failure → no orphan.","Negative test","Implemented")
add("REV-04","Revenue & Cash","Application","Revenue","Missing/duplicate sales not detected (completeness).","Completeness",
    "Sequential, gapless document numbering for sales/payments (per type/day).","Prev","Automated","Per txn","Eng Lead","P13",
    "common/doc-number.service.ts","Inspect numbering scheme.","Sample sequence for gaps/duplicates.","Doc-no sequence","Implemented")
add("REV-05","Revenue & Cash","Application","Cash","Cash drawer shortages/skimming undetected.","Existence/Accuracy",
    "Till reconciliation: opening float, counted vs expected cash, variance + X/Z shift reports.","Det","Auto+Manual","Per shift","Store Ops Mgr","P10/P16",
    "payments.service.ts (openTill/closeTill/aggregateTill)","Inspect variance computation.","Sample shifts: Z-report variance reviewed & signed.","Z-reports","Implemented")
add("REV-06","Revenue & Cash","Application","Cash; Revenue","Refunds exceed original payment (leakage/fraud).","Validity",
    "Over-refund prevention — refund + prior refunds ≤ captured, evaluated under a payment-row lock.","Prev","Automated","Per txn","Eng Lead","P10",
    "payments.service.ts (refund, FOR UPDATE)","Inspect lock + cumulative check.","Concurrent-refund test cannot exceed captured.","Refund test","Implemented")
add("REV-07","Revenue & Cash","Application","Revenue; Inventory","Refund posts but stock/GL not reversed (partial state).","Completeness/Accuracy",
    "Returns processed atomically — refund + restock + return record + GL reversal in one transaction.","Prev","Automated","Per txn","Eng Lead","P10",
    "returns.service.ts (createReturn)","Inspect single-transaction boundary.","Inject mid-flow failure → full rollback.","Atomicity test","Implemented")
add("REV-08","Revenue & Cash","Application","Accounts Receivable","Orders accepted beyond customer credit limit; credit holds set/lifted by one person; limit changes untraceable.","Valuation",
    "Credit-limit check at order — outstanding AR + order ≤ limit, under tenant-row lock; credit-hold block. Credit Manager may place/release manual holds and change limits, but a hold cannot be released by the user who placed it (SOD_SELF_RELEASE — requires an approver); every hold/release/limit-change is written to a credit-events audit trail.","Prev","Automated","Per order / per credit action","Controller","P10",
    "pos.service.ts (createOrder, FOR UPDATE); collections.service.ts (placeHold/releaseHold/changeLimit, creditEvents); cutover/compliance.ts + basics.ts (ToE)","Inspect AR sum + lock + limit logic; inspect hold/release SoD and credit-events logging.","Sample: outstanding+order ≤ limit allowed, > limit → CREDIT_LIMIT, credit-hold → CREDIT_HOLD; self-release → SOD_SELF_RELEASE, second person releases; limit change recorded old→new in credit-events (re-performed by the harness).","Credit test","Implemented")
add("REV-09","Revenue & Cash","Application","Cash","Forged payment-gateway callback flips a payment to captured.","Occurrence",
    "PSP webhook HMAC-SHA256 signature over raw body; fail-closed in production; out-of-band status re-verify.","Prev","Automated","Per callback","Eng Lead","P10/P11",
    "pos-terminal.controller.ts (PspWebhookController); crypto.verifyWebhookSignature","Inspect signature check + prod gate.","Replay/forged signature → 401; valid → accepted.","Webhook tests","Implemented")
add("REV-10","Revenue & Cash","Application","Revenue; Tax","Sales not posted to GL or posted unbalanced.","Completeness/Accuracy",
    "Sale/accept posts a balanced revenue + VAT journal entry to the GL automatically.","Auto","Automated","Per txn","Controller","P10",
    "ledger.service.ts (postEntry) from POS/CPQ flows","Walkthrough sale → JE.","Tie sample sales to GL postings.","Sale→GL tie-out","Implemented")
add("REV-11","Revenue & Cash","Application","Cash","Card settlements not reconciled to PSP payouts.","Existence/Accuracy",
    "Payment intents batched into settlement; settlement reconcile step.","Det","Auto+Manual","Daily","Controller","P16",
    "pos-terminal.service.ts (settle / reconcile)","Inspect settlement batching.","Sample batch reconciled to PSP statement.","Settlement recon","Implemented")
add("REV-12","Revenue & Cash","Application","Accounts Receivable","Overdue receivables not pursued; collection lapses; further credit extended to a defaulting customer.","Valuation/Authorization",
    "AR collections worklist + escalating dunning ladder (reminder→legal) recorded per invoice (DUN-), with the per-stage notice DISPATCHED to the customer (email/SMS/LINE via the messaging gateway; delivery outcome + recipient logged); credit-status / credit-check hold (over-limit OR 90+ days overdue). The serious-overdue hold is ENFORCED at POS/portal order entry (CREDIT_OVERDUE) under the FOR-UPDATE tenant lock, single-sourced with the collections on_hold threshold so the two never drift.","Det/Prev","Auto+Manual","Per overdue invoice / per order","AR / Controller","P10/P16",
    "collections.service.ts (worklist, recordDunning+dispatch, runDunningSweep, creditStatus, creditCheck, SERIOUS_OVERDUE_DAYS); messaging.service.ts (notice delivery); pos.service.ts (createOrder CREDIT_OVERDUE); collections.controller.ts (incl. cron-callable /collections/sweep); bi.service.ts (ar_collections_dunning scheduled job, daily via runDue); /finance Collections UI; cutover/basics.ts (ToE)","Inspect aging→stage logic, notice dispatch, hold decision, automated+scheduled sweep + order-entry gate.","Worklist ages open AR; manual + automated (idempotent) + daily-scheduled dunning recorded AND the notice delivered (message_log, channel auto-picked from contact); held customer's credit-check denied; a 90+ defaulter is blocked at order entry while a good-standing customer orders (re-performed by the harness).","Dunning log + message_log; sweep/scheduler run log; credit-hold report; CREDIT_OVERDUE rejections","Implemented")

# ---- Expenditure / Procurement / AP ----
add("EXP-01","Expenditure","Application","Accounts Payable; Inventory","Pay supplier for goods not ordered/received or wrong price.","Occurrence/Accuracy",
    "3-way match (PO ↔ GR ↔ Invoice) within tolerance gates AP payment.","Prev","Automated","Per invoice","AP / Controller","P10",
    "match/three-way-match.service.ts","Inspect match + tolerance + payment gate.","Sample invoices: payment blocked until matched.","Match results","Implemented")
add("EXP-02","Expenditure","Application","Accounts Payable","Payment to blocklisted/unapproved vendor.","Validity",
    "Supplier approval-status / blocklist check blocks PO & payment to blocked vendors.","Prev","Automated","Per txn","Procurement / AP","P10",
    "procurement.service.ts (SUPPLIER_BLOCKED)","Inspect vendor-status gate.","Attempt PO to blocked vendor → 422.","Vendor gate test","Implemented")
add("EXP-03","Expenditure","Application","Accounts Payable","PR/PO raised without authorization.","Authorization",
    "Purchase requisition / PO approval workflow (maker-checker) per approval matrix.","Prev","Auto+Manual","Per PR/PO","Procurement Mgr","P10",
    "workflow / approvals; procurement.service.ts","Inspect approval routing + thresholds.","Sample PRs: approved by authorized approver ≠ requester.","Approval trail","Partial")
add("EXP-04","Expenditure","Application","Accounts Payable","Match tolerance loosened to force payment.","Validity",
    "Match-tolerance configuration change restricted to 'creditors' permission.","Prev","Automated","Per change","Controller","P10/P11",
    "match.controller.ts (PUT tolerance @Permissions creditors)","Inspect permission on tolerance change.","Attempt change w/o perm → 403; log changes.","Config-change log","Implemented")
add("EXP-05","Expenditure","Application","Operating Expense; Accounts Payable","Employee/maintenance spend recorded in GL but not as a payable (escapes AP aging; AP sub-ledger ≠ GL control).","Completeness/Accuracy",
    "ESS expense approval and EAM work-order completion raise an AP payable via createApTxn (Dr 5100/5710 / Cr 2000) — not a bare GL post — so the liability appears in AP aging, settles through the AP pay flow, and keeps the AP sub-ledger ↔ GL 2000 reconciled (REC-01). Reimbursements vat-exempt; SoD approver ≠ claimant retained.","Prev","Automated","Per claim/WO","AP / Controller","P10",
    "ess.service.ts (approveExpense → createApTxn); eam.service.ts (updateWorkOrderStatus → createApTxn, acct 5710); finance.service.ts (createApTxn expense_account/tenant_id); cutover/ess.ts + basics.ts (ToE)","Inspect approval→AP routing + reconciliation.","Approve a claim / complete a WO → AP payable raised, paid via AP, sub-ledger ties to GL 2000 (re-performed by the harness).","AP sub-ledger; reimbursement/maintenance payables","Implemented")
add("EXP-06","Expenditure","Application","Accounts Payable; Cash","Vendor payment disbursed without independent approval — one person both books and pays a bill.","Authorization",
    "AP disbursement maker-checker — a payment is REQUESTED by a 'creditors' holder and APPROVED by a DIFFERENT user (approvals/gl_close); the bill's paid_amount and the cash-disbursement GL move only on approval. Booking a bill pre-paid in one call is blocked.","Prev","Auto+Manual","Per payment","AP / Controller","P10",
    "finance.service.ts (requestApPayment / approveApPayment requester≠approver / rejectApPayment); finance.controller.ts (creditors requests, approvals|gl_close approves); schema/finance.ts (ap_payments) + migration 0115; cutover/compliance.ts (ToE)","Inspect AP payment approval routing + SoD guard.","Sample payments: approver ≠ requester; no cash/GL effect until approved; self-approval → 403 SOD_VIOLATION; pre-paid creation → 400 (re-performed by the harness).","AP payment approvals","Implemented")
add("FA-06","Fixed Assets","Application","Property, Plant & Equipment","Equipment not maintained; maintenance uncontrolled; preventive maintenance missed; maintenance cost mis-stated or not tracked per asset.","Existence/Valuation",
    "EAM maintenance work orders against the asset register with a guarded lifecycle (open→in_progress→completed/cancelled; BAD_TRANSITION); preventive-maintenance schedules (time/meter) with an idempotent due-generation sweep (cron / daily scheduled job eam_pm_generate). Work-order labor/parts cost lines roll up to the WO actual cost (so the AP posting reflects real spend, not the estimate); per-asset reliability & cost KPIs (corrective failures, MTBF, downtime, total maintenance spend) support maintenance budgeting and repair-vs-replace decisions.","Det","Automated","Per WO / per sweep","Maintenance / FaAccountant","P13/P16",
    "eam.service.ts (work orders, PM schedules, runPmDue, meters, WO cost lines, reliability); eam.controller.ts; bi.service.ts (eam_pm_generate); cutover/basics.ts (ToE)","Inspect WO lifecycle guard + PM due logic; inspect cost-line roll-up into actual cost and reliability KPI computation.","Raise/complete a WO; PM sweep raises due preventive WOs (time + meter) and is idempotent; cost lines roll up to actual cost and drive the AP posting; reliability KPIs computed (re-performed by the harness).","Work-order log; PM schedule + sweep run log; WO cost lines; reliability KPI report","Implemented")

# ---- Inventory / COGS ----
add("INV-01","Inventory & COGS","Application","Inventory; COGS","Two terminals sell last unit → negative/oversold stock.","Existence/Valuation",
    "Bin-stock decrement under FOR UPDATE inside a transaction — concurrent picks serialize (no oversell).","Prev","Automated","Per txn","Eng Lead / Warehouse","P10",
    "wms.service.ts (pick)","Inspect lock + sufficiency check in tx.","Concurrent last-unit test → one PICK_SHORT.","Concurrency test","Implemented")
add("INV-02","Inventory & COGS","Application","Inventory","Stock movements not recorded (completeness of perpetual stock).","Completeness",
    "Perpetual stock-movement + lot ledger logging on every issue/receipt/return.","Det","Automated","Per txn","Warehouse","P13",
    "stock-ops; custStockLog; lotLedger","Inspect movement logging.","Tie sample movements to balances.","Stock ledger","Implemented")
add("INV-03","Inventory & COGS","Application","COGS","COGS misstated; consumption not costed.","Accuracy",
    "Inventory costing → COGS posting on consumption; recipe/BOM deduction + reversal on return.","Auto","Automated","Per txn","Controller","P10",
    "costing; menu/recipe.service.ts","Walkthrough consumption → COGS JE.","Tie sample consumption to COGS.","COGS tie-out","Implemented")
add("INV-04","Inventory & COGS","Application","Inventory","Book vs physical stock diverges; no count control.","Existence",
    "Periodic stocktake / cycle count; variance posted with review & approval.","Det","Auto+Manual","Periodic","Warehouse Mgr","P16",
    "stocktake (web); stock-ops","Inspect count + variance process.","Sample counts: variance reviewed & approved.","Count sheets","Partial")

# ---- General Ledger & Close ----
add("GL-01","General Ledger","Application","All","Unbalanced / one-sided journal entries.","Accuracy",
    "Double-entry balanced-by-construction — Σdebit=Σcredit enforced; each line single-sided & non-negative.","Prev","Automated","Per JE","Controller","P10",
    "ledger.service.ts (postEntry — UNBALANCED / INVALID_LINE)","Inspect balance + line invariants.","Attempt unbalanced JE → rejected.","Invariant test","Implemented")
add("GL-02","General Ledger","Application","All","Postings to a closed period distort reported results.","Cutoff",
    "Period-close lockout — postings to a CLOSED fiscal period are rejected (per-tenant calendar).","Prev","Automated","Per JE","Controller","P10",
    "ledger.service.ts (postEntry PERIOD_CLOSED); fiscal_periods","Inspect close check.","Post to closed period → rejected; review exceptions.","Close-lock test","Implemented")
add("GL-03","General Ledger","Application","Equity","Unauthorized year-end close / RE roll.","Cutoff/Authorization",
    "Year-end close restricted to 'exec'; closing entries roll to retained earnings.","Prev","Auto+Manual","Annual","CFO / Controller","P10/P11",
    "ledger.controller.ts (close-year @Permissions exec); closeYear","Inspect permission + closing logic.","Attempt close w/o exec → 403; review annual close.","Close package","Implemented")
add("GL-04","General Ledger","Application","All","Concurrent posting double-books the same source doc.","Completeness/Accuracy",
    "Ledger idempotency — UNIQUE (tenant,source,source_ref,ledger) + ON CONFLICT DO NOTHING in postEntry.","Prev","Automated","Per JE","Eng Lead / Controller","P10",
    "schema/ledger.ts (ux_je_idem); ledger.service.ts; migration 0058","Inspect unique index + conflict handling.","Concurrent identical post → single entry.","Dedup test","Implemented")
add("GL-05","General Ledger","Application","All","Manual JE posted without independent review (override risk).","Authorization",
    "Manual journal-entry maker-checker — preparer ≠ approver before a manual JE posts.","Prev","Auto+Manual","Per manual JE","Controller","P10",
    "ledger.service.ts (postEntry pendingApproval→Draft; approveEntry preparer≠approver; rejectEntry); ledger.controller.ts (gl_post posts / gl_close approves); cutover/compliance.ts (ToE)","Inspect JE approval routing.","Sample manual JEs: approver ≠ preparer; Draft excluded from balances until approved (re-performed by the harness).","JE approvals","Implemented")
add("GL-06","General Ledger","Application","All","Operator mis-posts to another tenant's books.","Validity",
    "HQ cross-tenant posting gated to Admin (explicit tenant override); others pinned to context (also RLS).","Prev","Automated","Per JE","Eng Lead","P10/P11",
    "ledger.controller.ts (hqTenant)","Inspect Admin gating of tenant_id.","Non-Admin tenant override ignored; Admin audited.","Override test","Implemented")
add("GL-07","General Ledger","Application","Cash","Statement of cash flows mis-stated or doesn't tie to the change in cash.","Accuracy/Completeness",
    "Statement of Cash Flows (indirect) reconstructed from the GL: net income + non-cash add-backs + working-capital + investing + financing; year-end CLOSE entries excluded; reconciles to Δcash (1000/1010/1020) by construction with a `reconciled` tie-out flag. A direct-method view presents the same operating cash by receipt/payment nature (receipts from customers, payments to suppliers, tax & payroll) and also reconciles to Δcash; a forward cash-flow forecast projects open AR (inflows) and AP (outflows) by due date from today's cash balance for treasury/collections planning.","Det","Automated","Per period","Controller","P10",
    "ledger.service.ts (cashFlowStatement, cashFlowDirect, cashFlowForecast); ledger.controller.ts (GET cash-flow, cash-flow-direct, cash-flow-forecast); cutover/basics.ts (ToE)","Inspect SCF derivation + reconciliation flag; inspect direct-method bucketing and forecast projection.","Run cash-flow over a seeded period — activities tie to the movement in cash; CLOSE excluded; direct method ties to the same operating cash and Δcash; forecast buckets open AR/AP by due week (re-performed by the harness).","Cash-flow reconciliation; direct-method buckets; forecast schedule","Implemented")
add("GL-08","General Ledger","Application","All","Standing/period accruals missed, posted unbalanced, or posted without approval; manual re-keying error.","Completeness/Accuracy/Authorization",
    "Recurring/template journal entries: a template's lines are validated balanced (Σdebit=Σcredit) AT SAVE TIME (unbalanced → UNBALANCED), so a broken template can never be persisted. The scheduled job gl_recurring_journals (cron / daily scheduler) posts every due template as a DRAFT JE through the normal maker-checker flow (GL-05) — a different user must approve before it affects balances — and rolls next_run_date forward. Idempotent: next_run_date advanced on posting + the (tenant,source,source_ref,ledger) idempotency key dedupes a same-day re-run. Templates can be paused/resumed without losing history.","Prev","Automated","Per due date / per template","Controller","P10",
    "ledger.service.ts (createRecurring balanced-at-save, runDueRecurring→postEntry pendingApproval, setRecurringActive); ledger.controller.ts (gl_post creates/runs); bi.service.ts (gl_recurring_journals); schema/ledger.ts (recurring_journals) + migration 0119; cutover/basics.ts (ToE)","Inspect template balance validation + scheduled posting as Draft + idempotency.","Unbalanced template → UNBALANCED; due template posts a Draft JE (excluded from TB) into the pending queue; re-run posts nothing; a second user approves → accrual hits the GL (re-performed by the harness).","Recurring-journal templates; scheduled run log; JE approvals","Implemented")
add("GL-09","General Ledger","Application","Prepaid expenses; Operating expense","Prepaid expense not amortized over its term — expense mis-stated, prepaid asset overstated.","Accuracy/Completeness",
    "Prepaid amortization schedules: a prepaid asset (insurance/rent paid up front) is registered once with a total + term; the scheduled job gl_prepaid_amortize posts a straight-line slice each period (Dr expense / Cr 1280), the last period taking the remainder so the asset fully clears. Idempotent per (schedule, period) via the JE idempotency key + next_run_date advance.","Det","Automated","Per period / per schedule","Controller","P10",
    "ledger.service.ts (createPrepaid, runDuePrepaid); ledger.controller.ts (gl_post); bi.service.ts (gl_prepaid_amortize); schema/ledger.ts (prepaid_schedules) + migration 0120; cutover/basics.ts (ToE)","Inspect amortization slice + final-period remainder + idempotency.","Register a 12-month prepaid; run amortizes 1/12 to expense; re-run posts nothing (re-performed by the harness).","Prepaid schedules; amortization run log","Implemented")
add("EXP-07","Expenditure","Application","Cash; Operating expense","Petty-cash / employee advance issued but never accounted for — cash leakage, expense unrecorded.","Existence/Completeness",
    "Employee cash advances: issuing an advance debits the 1180 Employee Advances control (Dr 1180 / Cr 1000); settlement clears it against actual spend + returned cash (Dr expense + Dr 1000 / Cr 1180), enforced to reconcile (settled_expense + returned_cash must equal the advance → SETTLE_MISMATCH). The 1180 balance is the outstanding float subject to review.","Det","Automated","Per advance","AP / Controller","P10",
    "finance.service.ts (issueAdvance, settleAdvance reconcile-or-reject, listAdvances outstanding); finance.controller.ts (creditors/exec); schema/finance.ts (employee_advances) + migration 0120; cutover/basics.ts (ToE)","Inspect issue/settle GL + settlement reconciliation guard.","Issue an advance (1180 up); settle 700+300 clears 1180; mismatched settle → SETTLE_MISMATCH (re-performed by the harness).","Advance register; outstanding float","Implemented")
add("FA-07","Fixed Assets","Application","Property, Plant & Equipment; Equity","Carrying amount not adjusted for revaluation/impairment — asset over/understated.","Valuation",
    "Asset revaluation / impairment: an upward revaluation credits the revaluation surplus in equity (Dr 1500 / Cr 3200); a downward revaluation (impairment) debits impairment loss (Dr 5820 / Cr 1500). The gross 1500 moves by the delta so the register stays tied to the GL; a no-op revaluation is rejected (NO_CHANGE). Every event is logged for the audit trail.","Det","Automated","Per event","FaAccountant / Controller","P10",
    "assets.service.ts (revalue up→3200 / down→5820, listRevaluations); assets.controller.ts (exec/creditors); schema/assets.ts (asset_revaluations) + migration 0120; cutover/basics.ts (ToE)","Inspect revaluation/impairment routing + audit log.","Upward revaluation credits 3200; impairment debits 5820; no-change → NO_CHANGE; both events logged (re-performed by the harness).","Revaluation log","Implemented")
add("LSE-01","Leases","Application","Right-of-use assets; Lease liabilities","Lease not capitalised (IFRS 16 / TFRS 16) — ROU asset + lease liability omitted; interest/depreciation mis-stated.","Completeness/Accuracy/Valuation",
    "Lease accounting: at commencement a right-of-use asset and lease liability are recognised at the present value of the lease payments (Dr 1600 / Cr 2600). The scheduled job lease_periodic_run posts each period — interest unwinding on the liability (Dr 5900), the cash payment reducing the liability (Dr 2600 / Cr 1000), and straight-line ROU depreciation (Dr 5210 / Cr 1690) — the last period clearing the liability + ROU exactly. Idempotent per (lease, period).","Det","Automated","Per period / per lease","Controller","P10",
    "leases.service.ts (createLease PV recognition, runDueLeases interest+payment+depreciation); leases.controller.ts (gl_post); bi.service.ts (lease_periodic_run); schema/leases.ts (leases) + migration 0120; cutover/basics.ts (ToE)","Inspect PV recognition + periodic interest/payment/depreciation + final-period clearing.","Commencement recognises ROU=liability=PV; periodic run posts interest+principal(=payment)+ROU depreciation; liability + ROU reduce (re-performed by the harness).","Lease register; periodic run log","Implemented")

# ---- Reconciliation ----
add("REC-01","Reconciliation","Application","All","Subledgers diverge from GL undetected.","Completeness/Accuracy",
    "Subledger-to-GL reconciliation per period: import GL, auto-match, certify (sign-off via 'approvals').","Det","Auto+Manual","Monthly","Controller","P16",
    "reconciliation.service.ts (importGlItems/autoMatch/certify)","Inspect recon + certify gate.","Sample periods: reconciled, unmatched cleared, certified.","Certified recon","Implemented")
add("REC-02","Reconciliation","Application","Cash","Bank balance not reconciled to GL.","Existence/Completeness",
    "Bank reconciliation against statements.","Det","Auto+Manual","Monthly","Controller","P16",
    "bank module; drizzle/0015_bank_reconciliation","Inspect bank-rec process.","Sample months reconciled & reviewed.","Bank rec","Implemented")
add("REC-03","Reconciliation","Application","Consolidation","Intercompany balances not eliminated/agreed.","Accuracy",
    "Intercompany reconciliation & elimination on consolidation.","Det","Auto+Manual","Period","Group Controller","P16",
    "intercompany; consolidation","Inspect IC matching + elimination.","Sample IC pairs agree & eliminate.","IC recon","Partial")

# ---- Tax ----
add("TAX-01","Tax","Application","Tax (VAT)","VAT computed incorrectly → filing/penalty risk.","Accuracy",
    "VAT via pluggable provider (TH 7%); unit-tested; no hard-coded rate.","Auto","Automated","Per txn","Tax / Controller","P10",
    "tax/tax-providers.ts; tax.service.ts; test/unit.test.ts","Inspect rate provider + tests.","Re-perform VAT on sample; tie to return.","VAT tests","Implemented")
add("TAX-02","Tax","Application","Tax; Revenue","Non-compliant e-Tax invoice / not transmitted to ETDA.","Accuracy/Compliance",
    "e-Tax invoice (ETDA UBL 2.1 XML) generation + email with CC to ETDA timestamp mailbox.","Auto","Automated","Per invoice","Tax","P10/P13",
    "tax-docs/etax-xml.ts; etax-email.service.ts (tested)","Inspect XML schema + email composer.","Sample invoices: valid XML, correct CC/escaping.","e-Tax samples","Implemented")
add("TAX-03","Tax","Application","Tax (WHT)","Withholding tax mis-computed / not reported.","Accuracy",
    "WHT computation and ภ.ง.ด. reporting.","Auto","Automated","Per txn / monthly","Tax","P10",
    "tax-reports; payroll","Inspect WHT logic + report.","Re-perform WHT on sample; tie to filing.","WHT report","Partial")

# ---- Payroll ----
add("PAY-01","Payroll","Application","Payroll expense/liability","Payroll, SSO and PIT mis-computed.","Accuracy",
    "Payroll engine — SSO 5% (cap 750), progressive PIT, payslip net; unit-tested.","Auto","Automated","Per run","HR / Payroll","P10",
    "payroll/payroll-calc.ts; test/unit.test.ts","Inspect calc + tests.","Re-perform sample payslips.","Payslip tests","Implemented")
add("PAY-02","Payroll","Application","Payroll liability","Statutory items (PF/OT/leave/ภ.ง.ด.1ก) wrong.","Accuracy/Compliance",
    "Provident fund, overtime, leave accrual and ภ.ง.ด.1ก reporting.","Auto","Automated","Per run / monthly","HR / Payroll","P10",
    "hcm; payroll (Phase 19)","Inspect statutory logic.","Sample run tied to statutory filings.","Filings","Partial")

# ---- Consolidation / FX ----
add("CON-01","Consolidation & FX","Application","Consolidation","Group consolidation mis-stated (ownership/currency).","Accuracy",
    "Consolidation run (ownership %, entity currency) gated by 'approvals'.","Auto+Manual","Period","Period","Group Controller","P10",
    "consolidation.service.ts (runConsolidation @Permissions approvals)","Inspect consolidation logic + gate.","Sample period: consolidated TB ties to entities.","Consol TB","Implemented")
add("CON-02","Consolidation & FX","Application","FX gain/loss","FX exposures not revalued at period-end.","Valuation",
    "FX revaluation at period-end rates; unrealized FX posted (acct 5400).","Auto+Manual","Period-end","Period","Controller","P10",
    "fx.service.ts (revalue / unrealized report)","Inspect reval logic + rates.","Recompute reval on sample balances.","FX reval JE","Implemented")

# write RCM sheet
rcm = wb.create_sheet("RCM")
rcm.sheet_view.showGridLines = False
for j,h in enumerate(HEAD, start=1):
    c = rcm.cell(row=1, column=j, value=h); c.font = f(10, True, "FFFFFF"); c.fill = fill(NAVY); c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); c.border = BORDER
    rcm.column_dimensions[get_column_letter(j)].width = W[j-1]
status_style = {"Implemented":(GREEN_F,GREEN_T),"Partial":(AMBER_F,AMBER_T),"Gap":(RED_F,RED_T)}
cycle_band = {}
for i,row in enumerate(R, start=2):
    band = BAND if (i % 2 == 0) else "FFFFFF"
    for j,val in enumerate(row, start=1):
        c = rcm.cell(row=i, column=j, value=val); c.font = f(9); c.alignment = WRAP; c.border = BORDER; c.fill = fill(band)
        if j in (1,3,8,9,10,12,17): c.alignment = WRAPC
    sc = rcm.cell(row=i, column=17); fc,tc = status_style.get(row[16], ("FFFFFF","000000")); sc.fill = fill(fc); sc.font = f(9, True, tc)
    rcm.cell(row=i, column=1).font = f(9, True, NAVY)
rcm.freeze_panes = "C2"
rcm.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{len(R)+1}"
rcm.row_dimensions[1].height = 34

# ============================================================== GAP REMEDIATION
gh = ["Control ID","Cycle","Gap / Weakness","Remediation Action","Owner","Phase","Target (relative)","Priority"]
gw = [11,17,34,46,16,10,16,11]
# phase mapping per earlier plan; priority H/M
# NOTE: Phase A (2026-06-23) closed AC-12, AC-13, CM-01, CM-02, CM-03, CM-04, CM-05, OP-01, OP-03 —
# removed from this remediation backlog (now Implemented in the Controls tab; one-time console [setup]
# steps tracked in docs/11-next-upgrade-realworld-roadmap.md and the audit-readiness plan).
GAP = [
 ("ITGC-OP-02","ITGC · Operations","DR/BCP plan not authored (RTO/RPO now defined in backup runbook).","Author full DR/BCP; schedule annual test.","CTO / DevOps","Phase 2","Month 3-5","Medium"),
 ("ITGC-OP-04","ITGC · Operations","Batch-job failure alerting not yet wired (alert spec authored).","Wire pg-boss job failure alerting + review evidence.","DevOps / Controller","Phase 2","Month 3","Medium"),
 ("ITGC-SD-01","ITGC · SDLC","No formal SDLC policy.","Author SDLC policy with design/test/UAT/go-live sign-offs.","Head of Eng / Product","Phase 2","Month 3-4","Medium"),
 ("ITGC-SD-02","ITGC · SDLC","Cutover reconciliation not documented.","Document source→target balance tie-out + sign-off for any migration.","Controller / Eng","Phase 2","Month 4","Medium"),
 ("ITGC-SD-03","ITGC · SDLC","Test coverage of key controls partial.","Expand key-control regression tests; CI archives dated evidence.","Head of Eng","Phase 2","Month 3-5","Medium"),
 ("EXP-03","Expenditure","PR/PO approval workflow partial.","Finalize PR/PO maker-checker against DoA thresholds.","Procurement Mgr","Phase 2","Month 3-4","Medium"),
 ("INV-04","Inventory & COGS","Stocktake variance review informal.","Formalize count cadence + variance review/approval sign-off.","Warehouse Mgr","Phase 2","Month 4","Medium"),
 ("REC-03","Reconciliation","Intercompany recon partial.","Formalize IC matching + elimination sign-off each period.","Group Controller","Phase 2-3","Month 4-6","Medium"),
 ("ITGC-AC-07","ITGC · Access","Login brute-force lockout + JWT revocation outstanding.","DONE: token → httpOnly cookie + CSRF + web CSP. Remaining: per-account login lockout/throttle + JWT jti revocation list.","IT Security / Eng","Phase 2","Month 2-3","Low"),
 ("TAX-03","Tax","WHT reporting partial.","Complete WHT calc coverage + monthly ภ.ง.ด. tie-out.","Tax","Phase 2-3","Month 4-6","Medium"),
 ("PAY-02","Payroll","Statutory payroll items partial.","Complete PF/OT/leave + ภ.ง.ด.1ก reconciliation to filings.","HR / Payroll","Phase 2-3","Month 4-6","Medium"),
 ("ELC-01","Entity-Level","No ethics policy / acknowledgements.","Issue code of conduct; annual acknowledgement register.","CEO / HR","Phase 1","Month 1-2","High"),
 ("ELC-02","Entity-Level","No audit-committee ICFR oversight.","Charter audit committee; quarterly ICFR review minutes.","Board","Phase 1","Month 1-2","High"),
 ("ELC-03","Entity-Level","No documented delegation of authority.","Author DoA/approval matrix; reconcile to RBAC roles.","CFO","Phase 1","Month 2","High"),
 ("ELC-04","Entity-Level","No whistleblower channel.","Stand up anonymous hotline + non-retaliation policy.","Audit Cttee","Phase 1","Month 2","Medium"),
 ("ELC-05","Entity-Level","No fraud risk assessment.","Perform annual fraud risk assessment; map to controls.","CFO / IA","Phase 2","Month 3","Medium"),
]
gap = wb.create_sheet("Gap Remediation")
gap.sheet_view.showGridLines = False
gap["A1"] = "Remediation Plan — Gap & Partial controls (sequenced to the readiness timeline)"; gap["A1"].font = f(12, True, NAVY)
gap.merge_cells("A1:H1")
for j,h in enumerate(gh, start=1):
    c = gap.cell(row=2, column=j, value=h); c.font = f(10, True, "FFFFFF"); c.fill = fill(STEEL); c.alignment = WRAPC; c.border = BORDER
    gap.column_dimensions[get_column_letter(j)].width = gw[j-1]
pri_fill = {"High":RED_F,"Medium":AMBER_F}
for i,row in enumerate(GAP, start=3):
    band = BAND if (i % 2 == 1) else "FFFFFF"
    for j,val in enumerate(row, start=1):
        c = gap.cell(row=i, column=j, value=val); c.font = f(9); c.alignment = WRAP; c.border = BORDER; c.fill = fill(band)
        if j in (1,2,5,6,7,8): c.alignment = WRAPC
    pc = gap.cell(row=i, column=8); pc.fill = fill(pri_fill.get(row[7],"FFFFFF")); pc.font = f(9, True)
    gap.cell(row=i, column=1).font = f(9, True, NAVY)
gap.freeze_panes = "A3"; gap.auto_filter.ref = f"A2:H{len(GAP)+2}"; gap.row_dimensions[2].height = 28

# ============================================================== COSO MAPPING
coso = wb.create_sheet("COSO Mapping")
coso.sheet_view.showGridLines = False
coso["A1"] = "COSO 2013 — 17 Principles coverage"; coso["A1"].font = f(12, True, NAVY); coso.merge_cells("A1:D1")
ch = ["Component","#","Principle","Covered by (examples)"]
cwd = [22,5,55,40]
for j,h in enumerate(ch, start=1):
    c = coso.cell(row=2, column=j, value=h); c.font = f(10, True, "FFFFFF"); c.fill = fill(STEEL); c.alignment = WRAPC; c.border = BORDER
    coso.column_dimensions[get_column_letter(j)].width = cwd[j-1]
COSO = [
 ("Control Environment",1,"Commitment to integrity & ethical values","ELC-01, ELC-04"),
 ("Control Environment",2,"Board exercises oversight responsibility","ELC-02"),
 ("Control Environment",3,"Establishes structure, authority & responsibility","ELC-03, ITGC-AC-02"),
 ("Control Environment",4,"Commitment to competence","HR / training (entity)"),
 ("Control Environment",5,"Enforces accountability","ELC-03, ITGC-AC-08/09"),
 ("Risk Assessment",6,"Specifies suitable objectives","Scoping memo (to author)"),
 ("Risk Assessment",7,"Identifies & analyzes risk","Scoping / risk assessment"),
 ("Risk Assessment",8,"Assesses fraud risk","ELC-05"),
 ("Risk Assessment",9,"Identifies & analyzes significant change","ITGC-CM-*, ITGC-SD-01"),
 ("Control Activities",10,"Selects & develops control activities","REV-*, EXP-*, INV-*, GL-*, REC-*"),
 ("Control Activities",11,"Selects & develops GENERAL CONTROLS over technology (ITGC)","All ITGC-AC/CM/SD/OP"),
 ("Control Activities",12,"Deploys through policies & procedures","SDLC/DoA/close policies (to author)"),
 ("Information & Comm.",13,"Uses relevant, quality information","REV-04, ITGC-AC-10/11, GL-*"),
 ("Information & Comm.",14,"Communicates internally","Close calendar / reporting"),
 ("Information & Comm.",15,"Communicates externally","ELC-04; investor/auditor comms"),
 ("Monitoring",16,"Ongoing &/or separate evaluations","REV-05, REC-01/02, ITGC-SD-03, ELC-06"),
 ("Monitoring",17,"Evaluates & communicates deficiencies","Deficiency log / 404(a) assessment"),
]
for i,row in enumerate(COSO, start=3):
    band = BAND if (i % 2 == 1) else "FFFFFF"
    for j,val in enumerate(row, start=1):
        c = coso.cell(row=i, column=j, value=val); c.font = f(9); c.alignment = WRAP; c.border = BORDER; c.fill = fill(band)
        if j==2: c.alignment = WRAPC
    if row[1] == 11:
        for j in range(1,5): coso.cell(row=i, column=j).font = f(9, True, NAVY)
coso.freeze_panes = "A3"

# ============================================================== DASHBOARD (formulas)
dash = wb.create_sheet("Summary")
dash.sheet_view.showGridLines = False
dash["B2"] = "ICFR Readiness — Control Summary"; dash["B2"].font = f(14, True, NAVY); dash.merge_cells("B2:E2")
last = len(R)+1
dash["B4"] = "By Status"; dash["B4"].font = f(11, True, STEEL)
sd = [("Implemented",GREEN_F,GREEN_T),("Partial",AMBER_F,AMBER_T),("Gap",RED_F,RED_T)]
for i,(s,fc,tc) in enumerate(sd):
    r=5+i
    dash[f"B{r}"]=s; dash[f"B{r}"].fill=fill(fc); dash[f"B{r}"].font=f(10,True,tc)
    dash[f"C{r}"]=f'=COUNTIF(RCM!$Q$2:$Q${last},B{r})'; dash[f"C{r}"].font=f(10)
dash["B8"]="Total"; dash["B8"].font=f(10,True)
dash["C8"]=f'=COUNTA(RCM!$A$2:$A${last})'; dash["C8"].font=f(10,True)
dash["B10"]="By Category"; dash["B10"].font=f(11,True,STEEL)
for i,cat in enumerate(["Entity","ITGC","Application"]):
    r=11+i
    dash[f"B{r}"]=cat; dash[f"B{r}"].font=f(10)
    dash[f"C{r}"]=f'=COUNTIF(RCM!$C$2:$C${last},B{r})'; dash[f"C{r}"].font=f(10)
dash["B15"]="% Implemented"; dash["B15"].font=f(10,True)
dash["C15"]='=IFERROR(C5/C8,0)'; dash["C15"].font=f(10,True); dash["C15"].number_format='0.0%'
dash["B17"]=("Read with the 'RCM' tab (full control detail + tests), 'Gap Remediation' (sequenced fixes), and 'COSO Mapping' "
             "(17-principle coverage). Status counts update automatically from the RCM.")
dash.merge_cells("B17:E19"); dash["B17"].font=f(9,False,GREY); dash["B17"].alignment=WRAP
for col,w in (("A",3),("B",22),("C",12),("D",3),("E",40)): dash.column_dimensions[col].width=w

# order tabs
wb.move_sheet("Cover", -wb.sheetnames.index("Cover"))
out = "compliance/Oshinei_ERP_SOX_RCM_v1.xlsx"
wb.save(out)
print("WROTE", out, "| controls:", len(R), "| gaps:", len(GAP))
