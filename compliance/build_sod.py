from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT="Arial"; DATE="2026-06-22"
NAVY="1F4E78"; STEEL="2E5C8A"; BAND="F2F6FB"; GREY="808080"
HI_F="FFC7CE"; HI_T="9C0006"; MD_F="FFEB9C"; MD_T="9C6500"; LO_F="C6EFCE"; LO_T="006100"
CONF_F="C00000"; CLEAR_F="E2EFDA"; INH_F="BFBFBF"

def f(sz=10,b=False,c="000000"): return Font(name=FONT,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top"); CTR=Alignment(wrap_text=True,vertical="center",horizontal="center")

# ---------- source of truth: permissions.ts ----------
PERMS=["pos","dashboard","order_mgt","claim_mgt","crm","users","warehouse","procurement",
"creditors","ar","delivery","returns","pricelist","lots","locations","promos","mobile",
"images","masterdata","bom_master","planner","exec","order_cust","cust_dash",
"cust_inventory","cust_pos","cust_bom","cust_variance","loyalty","survey",
"cust_my_crm","cust_my_suppliers","cust_my_pos","cust_my_users","marketing","track","ai_chat","approvals"]

ROLES={
 "Admin":set(PERMS),
 "Sales":{"pos","dashboard","exec","order_mgt","claim_mgt","crm","ar","delivery","returns","pricelist","promos","marketing","planner","approvals"},
 "Procurement":{"procurement","creditors","ar","delivery","masterdata","approvals"},
 "Planner":{"dashboard","exec","warehouse","procurement","planner","masterdata","approvals"},
 "Warehouse":{"warehouse","lots","locations","mobile","images","masterdata"},
 "Customer":{"order_cust","cust_pos","cust_dash","cust_inventory","cust_bom","cust_variance","loyalty","survey","track","cust_my_crm","cust_my_suppliers","cust_my_pos","cust_my_users"},
}
ROLE_ORDER=["Admin","Sales","Procurement","Planner","Warehouse","Customer"]

# group, description (duty granted), sensitivity, code/route reference
META={
 "pos":("Sales & Orders","Record POS sales; (bundled) issue refunds/voids & manage till","High","pos.controller; payments.controller"),
 "order_mgt":("Sales & Orders","Manage sales orders / order status","Medium","pos OrdersController"),
 "claim_mgt":("Sales & Orders","Manage claims","Medium","claims"),
 "crm":("Sales & Orders","Customer master / CRM / pipeline","High","crm; customers; pipeline"),
 "delivery":("Sales & Orders","Delivery dispatch","Medium","delivery"),
 "returns":("Sales & Orders","Process customer returns","High","returns.controller"),
 "pricelist":("Sales & Orders","Maintain price lists","High","pricing"),
 "promos":("Sales & Orders","Maintain promotions/discounts","High","marketing/promos"),
 "dashboard":("Dashboard & Analytics","View operational dashboards","Low","dashboard"),
 "exec":("Dashboard & Analytics","Executive/finance: GL postings, period & year-end close, recon prep","High","ledger.controller; reconciliation.controller"),
 "planner":("Dashboard & Analytics","Planning / forecasting","Medium","planning"),
 "marketing":("Dashboard & Analytics","Marketing campaigns","Low","marketing"),
 "warehouse":("Warehouse","Warehouse ops: receive, adjust stock, custody & counts","High","wms; stock-ops; inventory"),
 "lots":("Warehouse","Lot/batch management","Medium","lots"),
 "locations":("Warehouse","Location/bin management","Medium","wms locations"),
 "mobile":("Warehouse","Mobile scanning","Low","scan; mobile-scan"),
 "images":("Warehouse","Item images","Low","images"),
 "ar":("Finance & AR/AP","Accounts receivable; AR/JE posting","High","ledger.controller; finance"),
 "creditors":("Finance & AR/AP","Accounts payable: pay vendors, match tolerance","High","payments; match.controller; ledger"),
 "procurement":("Procurement","Raise purchase requisitions / POs","High","procurement.controller"),
 "masterdata":("Administration","Master data: vendor/item/config maintenance","High","masterdata; admin-config"),
 "bom_master":("Administration","BOM master maintenance","Medium","bom; mfg"),
 "users":("Administration","User & access administration (grant permissions)","High","admin-users; platform"),
 "ai_chat":("Administration","AI assistant","Low","ai"),
 "approvals":("Administration","Approve workflow items; certify reconciliation; run consolidation","High","workflow; reconciliation certify; consolidation"),
 "order_cust":("Customer Portal","Customer self-order","Low","portal order"),
 "cust_pos":("Customer Portal","Customer self-checkout (portal POS)","Low","portal/pos"),
 "cust_dash":("Customer Portal","Customer dashboard","Low","portal dashboard"),
 "cust_inventory":("Customer Portal","Customer inventory view","Low","portal inventory"),
 "cust_bom":("Customer Portal","Customer BOM view","Low","portal bom"),
 "cust_variance":("Customer Portal","Customer variance view","Low","portal variance"),
 "loyalty":("Customer Portal","Loyalty points (own)","Low","loyalty"),
 "survey":("Customer Portal","Surveys","Low","survey"),
 "track":("Customer Portal","Order tracking","Low","track"),
 "cust_my_crm":("My Business","Own-org CRM","Low","portal/my"),
 "cust_my_suppliers":("My Business","Own-org suppliers","Low","portal/my"),
 "cust_my_pos":("My Business","Own-org purchase orders","Low","portal/my"),
 "cust_my_users":("My Business","Own-org user management","Medium","portal/my/users"),
}
SENS_STYLE={"High":(HI_F,HI_T),"Medium":(MD_F,MD_T),"Low":(LO_F,LO_T)}

# ---------- SoD conflict rules: dutyA perms vs dutyB perms ----------
# A role conflicts when it holds >=1 perm from A AND >=1 from B (intra-perm rules use A==B).
RULES=[
 ("R01","Access administration","Any transactional duty",{"users"},
   {"pos","order_mgt","procurement","creditors","ar","returns","pricelist","promos","masterdata","warehouse","exec"},
   "User can grant/modify access (incl. own) and also transact — self-authorize and cover tracks.","High",
   "Isolate access admin to a dedicated, non-transacting role; log all permission changes; quarterly UAR."),
 ("R02","Maintain vendor/master data","Disburse AP / pay vendors",{"masterdata"},{"creditors"},
   "Create a fictitious/edited vendor and pay it — classic disbursement fraud.","High",
   "Separate vendor-master maintenance from AP payment; vendor-change report reviewed independently."),
 ("R03","Raise purchase requisition / PO","Approve & pay AP",{"procurement"},{"creditors"},
   "Originate a purchase and also pay it — unauthorized/duplicate spend.","High",
   "Split buying from paying; route PO/payment through maker-checker approvals."),
 ("R04","Purchase ordering","Goods receipt / warehouse custody",{"procurement"},{"warehouse"},
   "Order goods and confirm receipt — defeats 3-way match (phantom receipts).","High",
   "Separate procurement from receiving; independent GR; rely on 3-way match (EXP-01)."),
 ("R05","Post journal entries","Close fiscal period / year",{"exec","creditors","ar"},{"exec"},
   "Post entries and also close the period — conceal misstatement after the fact.","High",
   "Restrict period/year close to a finance approver distinct from JE preparers; JE maker-checker (GL-05)."),
 ("R06","Prepare reconciliation","Certify reconciliation",{"exec"},{"approvals"},
   "Prepare and self-certify a reconciliation — no independent review.","Medium",
   "Preparer must differ from certifier; certification by an independent approver."),
 ("R07","Initiate transactions","Approve workflow items",{"procurement","pos","ar","creditors","order_mgt"},{"approvals"},
   "Initiate a transaction and approve it in the workflow — self-approval.","High",
   "Approver must differ from initiator; enforce in the approval engine."),
 ("R08","Record sale","Issue refund / reconcile till (within 'pos')",{"pos"},{"pos"},
   "The 'pos' permission BUNDLES sell + refund/void + till close — one cashier can ring, refund and reconcile their own drawer.","High",
   "Split 'pos' into sub-permissions (sell vs refund/void vs till close); require manager auth for refund/void; independent till count."),
 ("R09","Maintain customer / credit master","Enter sales orders",{"crm","masterdata"},{"pos","order_mgt","order_cust"},
   "Raise a customer's credit limit and then sell on credit to them — bad-debt / collusion risk.","Medium",
   "Separate customer/credit-limit maintenance from order entry; credit-change report reviewed."),
 ("R10","Maintain prices / promotions","Enter sales",{"pricelist","promos"},{"pos","order_mgt"},
   "Set a price/discount and sell at it — under-pricing/collusion.","Medium",
   "Separate price/promo maintenance from selling; price-override report reviewed."),
 ("R11","Adjust inventory","Stock custody & counting (within 'warehouse')",{"warehouse","lots","locations"},{"warehouse"},
   "'warehouse' bundles adjust + custody + count — conceal shrink/theft via adjustments.","Medium",
   "Separate adjustment authority from physical count; independent count + variance approval (INV-04)."),
 ("R12","Process returns","Issue refund",{"returns"},{"pos","ar"},
   "Process a return and issue the matching refund unchecked.","Medium",
   "Independent approval of refunds on returns; over-return guard (REV-07) + detective review."),
 ("R13","Maintain master data / config","Transact on it",{"masterdata","bom_master"},{"pos","order_mgt","procurement","creditors","ar"},
   "Change configuration/master data and transact against it without review.","Medium",
   "Segregate config from operations; master-data change log reviewed independently."),
 # ---- CRM / loyalty single-duty conflicts (Phase 4) — loyalty points are a TFRS-15 liability, so issuing
 #      loyalty value must be segregated from using/creating it. Granular crm_* perms (not held by the coarse
 #      named roles), so these are clean in the to-be design; only the Admin superuser holds all sides. ----
 ("R14","Configure rewards / vouchers","POS redemption at till",{"crm_reward"},{"pos_sell"},
   "Create a reward/voucher and redeem it for oneself at the till.","High",
   "Separate reward-catalog configuration from POS redemption; review reward-change + redemption reports."),
 ("R15","Manual points adjustment","Member master maintenance",{"crm_points_adjust"},{"crm_member"},
   "Enrol a ghost member and credit points to it.","High",
   "Separate member enrolment from points adjustment; over-threshold adjust via maker-checker approval."),
 ("R16","Campaign issuance of point-bearing value","Points adjustment",{"crm_campaign"},{"crm_points_adjust"},
   "Self-issue loyalty value through two channels (campaign coupons + manual adjustment).","High",
   "Separate campaign issuance from points adjustment; review issuance + adjustment logs."),
]

def conflicts_for(role_perms):
    out=[]
    for rid,da,db,A,B,risk,sev,mit in RULES:
        ha=role_perms&A; hb=role_perms&B
        if ha and hb:
            held=sorted(ha|hb)
            out.append((rid,da,db,held,risk,sev,mit))
    return out

wb=Workbook()

# ================= COVER
cov=wb.active; cov.title="Cover"; cov.sheet_view.showGridLines=False
for col,w in (("A",3),("B",30),("C",95)): cov.column_dimensions[col].width=w
cov["B2"]="Segregation of Duties (SoD) — Conflict Matrix"; cov["B2"].font=f(20,True,NAVY)
cov["B3"]="Computed from the live RBAC role→permission model (37+1 permissions × 6 roles)"; cov["B3"].font=f(11,False,STEEL)
cov.merge_cells("B2:C2"); cov.merge_cells("B3:C3")
def crow(r,l,v):
    cov[f"B{r}"]=l; cov[f"B{r}"].font=f(10,True,NAVY); cov[f"B{r}"].alignment=WRAP
    cov[f"C{r}"]=v; cov[f"C{r}"].font=f(10); cov[f"C{r}"].alignment=WRAP
crow(6,"Entity","Invisible Consulting — Oshinei Enterprise ERP (EGC)")
crow(7,"Source of truth","packages/shared/src/permissions.ts — PERMISSIONS + DEFAULT_ROLE_PERMISSIONS")
crow(8,"Enforcement today","RBAC via @Permissions (OR-semantics) + PermissionsGuard; tenant isolation via RLS")
crow(9,"Method","16 SoD conflict rules; a role conflicts when it holds duties on BOTH sides of a rule")
crow(10,"Prepared (draft)",DATE+"  ·  Version 1.0 — review with auditor + SOX advisor")
cov["B13"]="How to read"; cov["B13"].font=f(11,True,NAVY)
cov["B14"]=("Tabs: Permission Inventory (what each permission grants + sensitivity) · SoD Rules (the conflict library) · "
 "Role × Permission (assignment grid) · Conflict Matrix (Role × Rule heat grid + counts) · Detected Conflicts (every "
 "role/rule hit, with offending permissions + mitigation).")
cov.merge_cells("B14:C16"); cov["B14"].font=f(10); cov["B14"].alignment=WRAP
cov["B18"]="Critical caveats (state these to the auditor)"; cov["B18"].font=f(11,True,NAVY)
cov["B19"]=("1) SoD is ultimately PER-USER. This baseline is at the ROLE level; users may carry per-user permission overrides "
 "(resolvePermissions userOverride) — actual user×permission assignments must also be reviewed (ties to ITGC-AC-08 UAR).\n"
 "2) Several permissions are COARSE and bundle conflicting duties inside one key (esp. 'pos' = sell+refund+till; 'warehouse' = "
 "adjust+custody+count). RBAC alone cannot separate these — they need sub-permissions or compensating detective controls.\n"
 "3) 'Admin' holds ALL permissions and therefore violates every rule by design. That is expected for a superuser, but makes "
 "Admin the HIGHEST-risk role: keep named Admins to a minimum, enforce MFA (ITGC-AC-06), and rely on the tamper-evident "
 "audit log + hash-chained journal (ITGC-AC-10/11) as compensating controls.\n"
 "4) @Permissions uses OR-semantics: holding ANY one listed permission grants the route — widening effective access.")
cov.merge_cells("B19:C27"); cov["B19"].font=f(10); cov["B19"].alignment=WRAP
cov["B29"]=("Disclaimer: working draft to accelerate audit readiness; not legal/accounting advice. Validate the rule library and "
 "role design with your independent auditor and SOX advisor.")
cov.merge_cells("B29:C31"); cov["B29"].font=f(9,False,GREY); cov["B29"].alignment=WRAP

# ================= PERMISSION INVENTORY
inv=wb.create_sheet("Permission Inventory"); inv.sheet_view.showGridLines=False
h=["Permission","Group","Duty / Capability granted","Financial sensitivity","Held by roles","Code / Route reference"]
ww=[18,22,46,16,30,34]
for j,x in enumerate(h,1):
    c=inv.cell(1,j,x); c.font=f(10,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=CTR; c.border=BORDER
    inv.column_dimensions[get_column_letter(j)].width=ww[j-1]
for i,p in enumerate(PERMS,2):
    grp,desc,sens,ref=META[p]
    held=", ".join(r for r in ROLE_ORDER if p in ROLES[r])
    band=BAND if i%2==0 else "FFFFFF"
    vals=[p,grp,desc,sens,held,ref]
    for j,v in enumerate(vals,1):
        c=inv.cell(i,j,v); c.font=f(9); c.alignment=WRAP; c.border=BORDER; c.fill=fill(band)
        if j in (1,4): c.alignment=CTR
    sc=inv.cell(i,4); fc,tc=SENS_STYLE[sens]; sc.fill=fill(fc); sc.font=f(9,True,tc)
    inv.cell(i,1).font=f(9,True,NAVY)
inv.freeze_panes="A2"; inv.auto_filter.ref=f"A1:F{len(PERMS)+1}"; inv.row_dimensions[1].height=30

# ================= SoD RULES
sr=wb.create_sheet("SoD Rules"); sr.sheet_view.showGridLines=False
h=["Rule","Duty A","Duty B","Permissions (A)","Permissions (B)","Risk if combined","Severity","Required mitigation"]
ww=[7,24,24,20,20,40,10,44]
for j,x in enumerate(h,1):
    c=sr.cell(1,j,x); c.font=f(10,True,"FFFFFF"); c.fill=fill(STEEL); c.alignment=CTR; c.border=BORDER
    sr.column_dimensions[get_column_letter(j)].width=ww[j-1]
for i,(rid,da,db,A,B,risk,sev,mit) in enumerate(RULES,2):
    band=BAND if i%2==0 else "FFFFFF"
    vals=[rid,da,db,", ".join(sorted(A)),", ".join(sorted(B)),risk,sev,mit]
    for j,v in enumerate(vals,1):
        c=sr.cell(i,j,v); c.font=f(9); c.alignment=WRAP; c.border=BORDER; c.fill=fill(band)
        if j in (1,7): c.alignment=CTR
    fc,tc=SENS_STYLE[sev]; sc=sr.cell(i,7); sc.fill=fill(fc); sc.font=f(9,True,tc)
    sr.cell(i,1).font=f(9,True,NAVY)
sr.freeze_panes="A2"; sr.auto_filter.ref=f"A1:H{len(RULES)+1}"; sr.row_dimensions[1].height=28

# ================= ROLE x PERMISSION
rp=wb.create_sheet("Role x Permission"); rp.sheet_view.showGridLines=False
rp.cell(1,1,"Permission"); rp.cell(1,1).font=f(10,True,"FFFFFF"); rp.cell(1,1).fill=fill(NAVY); rp.cell(1,1).border=BORDER; rp.cell(1,1).alignment=CTR
rp.cell(1,2,"Sensitivity"); rp.cell(1,2).font=f(10,True,"FFFFFF"); rp.cell(1,2).fill=fill(NAVY); rp.cell(1,2).border=BORDER; rp.cell(1,2).alignment=CTR
for j,role in enumerate(ROLE_ORDER,3):
    c=rp.cell(1,j,role); c.font=f(10,True,"FFFFFF"); c.fill=fill(NAVY); c.border=BORDER; c.alignment=CTR
    rp.column_dimensions[get_column_letter(j)].width=12
rp.column_dimensions["A"].width=18; rp.column_dimensions["B"].width=12
for i,p in enumerate(PERMS,2):
    sens=META[p][2]; band=BAND if i%2==0 else "FFFFFF"
    a=rp.cell(i,1,p); a.font=f(9,True,NAVY); a.border=BORDER; a.fill=fill(band); a.alignment=CTR
    s=rp.cell(i,2,sens); fc,tc=SENS_STYLE[sens]; s.fill=fill(fc); s.font=f(9,True,tc); s.border=BORDER; s.alignment=CTR
    for j,role in enumerate(ROLE_ORDER,3):
        has=p in ROLES[role]; c=rp.cell(i,j,"✕" if has else "")
        c.border=BORDER; c.alignment=CTR; c.font=f(10,True,"FFFFFF" if has else "000000")
        c.fill=fill("4472C4" if has else band)
cnt=len(PERMS)+2
rp.cell(cnt,1,"TOTAL"); rp.cell(cnt,1).font=f(10,True)
for j,role in enumerate(ROLE_ORDER,3):
    col=get_column_letter(j)
    c=rp.cell(cnt,j,f'=COUNTIF({col}2:{col}{len(PERMS)+1},"✕")'); c.font=f(10,True); c.alignment=CTR; c.border=BORDER
rp.freeze_panes="C2"; rp.row_dimensions[1].height=18

# ================= CONFLICT MATRIX (Role x Rule)
cm=wb.create_sheet("Conflict Matrix"); cm.sheet_view.showGridLines=False
cm.cell(1,1,"Role"); cm.cell(1,1).font=f(10,True,"FFFFFF"); cm.cell(1,1).fill=fill(NAVY); cm.cell(1,1).border=BORDER; cm.cell(1,1).alignment=CTR
cm.column_dimensions["A"].width=14
for j,(rid,*_ ) in enumerate(RULES,2):
    c=cm.cell(1,j,rid); c.font=f(9,True,"FFFFFF"); c.fill=fill(NAVY); c.border=BORDER; c.alignment=CTR
    cm.column_dimensions[get_column_letter(j)].width=6
ccol=len(RULES)+2
c=cm.cell(1,ccol,"# Conflicts"); c.font=f(9,True,"FFFFFF"); c.fill=fill(NAVY); c.border=BORDER; c.alignment=CTR
cm.column_dimensions[get_column_letter(ccol)].width=11
detected=[]
for i,role in enumerate(ROLE_ORDER,2):
    a=cm.cell(i,1,role); a.font=f(10,True,NAVY); a.border=BORDER; a.alignment=CTR
    rc=conflicts_for(ROLES[role]); hits={x[0] for x in rc}
    for x in rc: detected.append((role,)+x)
    n=0
    for j,(rid,*_ ) in enumerate(RULES,2):
        conf=rid in hits
        cell=cm.cell(i,j)
        if role=="Admin":
            cell.value="▲"; cell.fill=fill(INH_F); cell.font=f(9,True,"FFFFFF")
        elif conf:
            cell.value="✕"; cell.fill=fill(CONF_F); cell.font=f(9,True,"FFFFFF"); n+=1
        else:
            cell.value="·"; cell.fill=fill(CLEAR_F); cell.font=f(9,False,"006100")
        cell.border=BORDER; cell.alignment=CTR
    tot=cm.cell(i,ccol, len(RULES) if role=="Admin" else n); tot.font=f(10,True); tot.alignment=CTR; tot.border=BORDER
    if role=="Admin": tot.value="ALL (▲)"; tot.font=f(9,True,HI_T)
cm.cell(len(ROLE_ORDER)+3,1,"Legend:"); cm.cell(len(ROLE_ORDER)+3,1).font=f(9,True)
cm.cell(len(ROLE_ORDER)+4,1,"✕ conflict"); cm.cell(len(ROLE_ORDER)+4,1).fill=fill(CONF_F); cm.cell(len(ROLE_ORDER)+4,1).font=f(9,True,"FFFFFF"); cm.cell(len(ROLE_ORDER)+4,1).alignment=CTR
cm.cell(len(ROLE_ORDER)+4,2,"· clear"); cm.cell(len(ROLE_ORDER)+4,2).fill=fill(CLEAR_F); cm.cell(len(ROLE_ORDER)+4,2).font=f(9,False,"006100"); cm.cell(len(ROLE_ORDER)+4,2).alignment=CTR
cm.cell(len(ROLE_ORDER)+4,3,"▲ inherent (superuser)"); cm.cell(len(ROLE_ORDER)+4,3).fill=fill(INH_F); cm.cell(len(ROLE_ORDER)+4,3).font=f(9,True,"FFFFFF"); cm.cell(len(ROLE_ORDER)+4,3).alignment=CTR
cm.merge_cells(start_row=len(ROLE_ORDER)+4,start_column=3,end_row=len(ROLE_ORDER)+4,end_column=5)
cm.freeze_panes="B2"; cm.row_dimensions[1].height=16

# ================= DETECTED CONFLICTS
dc=wb.create_sheet("Detected Conflicts"); dc.sheet_view.showGridLines=False
h=["Role","Rule","Duty A","Duty B","Permissions held (offending)","Risk","Severity","Recommended mitigation"]
ww=[13,7,22,24,28,38,10,44]
for j,x in enumerate(h,1):
    c=dc.cell(1,j,x); c.font=f(10,True,"FFFFFF"); c.fill=fill(STEEL); c.alignment=CTR; c.border=BORDER
    dc.column_dimensions[get_column_letter(j)].width=ww[j-1]
# exclude Admin's full list to keep it actionable; add one Admin summary row
det_nonadmin=[d for d in detected if d[0]!="Admin"]
det_nonadmin.sort(key=lambda d:(0 if d[5]=="High" else 1, d[0]))
ri=2
for (role,rid,da,db,held,risk,sev,mit) in det_nonadmin:
    band=BAND if ri%2==0 else "FFFFFF"
    vals=[role,rid,da,db,", ".join(held),risk,sev,mit]
    for j,v in enumerate(vals,1):
        c=dc.cell(ri,j,v); c.font=f(9); c.alignment=WRAP; c.border=BORDER; c.fill=fill(band)
        if j in (1,2,7): c.alignment=CTR
    fc,tc=SENS_STYLE[sev]; sc=dc.cell(ri,7); sc.fill=fill(fc); sc.font=f(9,True,tc)
    dc.cell(ri,1).font=f(9,True,NAVY)
    ri+=1
# Admin summary row
for j,v in enumerate(["Admin","ALL","(superuser)","(superuser)","ALL 38 permissions",
   "Holds every duty — inherently violates all 16 rules by design.","High",
   "Minimize named Admins; MFA (ITGC-AC-06); break-glass procedure; full reliance on audit log + hash-chained journal (ITGC-AC-10/11) + periodic privileged-access review."],1):
    c=dc.cell(ri,j,v); c.font=f(9); c.alignment=WRAP; c.border=BORDER; c.fill=fill(INH_F if j<=5 else "FFFFFF")
    if j in (1,2,7): c.alignment=CTR
dc.cell(ri,1).font=f(9,True,"FFFFFF"); dc.cell(ri,7).fill=fill(HI_F); dc.cell(ri,7).font=f(9,True,HI_T)
dc.freeze_panes="A2"; dc.auto_filter.ref=f"A1:H{ri}"; dc.row_dimensions[1].height=28

# ================= PROPOSED (TO-BE) REMEDIATED DESIGN =================
# 1) split coarse permissions into single-duty sub-permissions
SPLITS=[
 ("pos","pos_sell · pos_refund · pos_till","Separate selling from refund/void and from till reconciliation.","R08, R12","High"),
 ("warehouse","wh_receive · wh_adjust · wh_count · wh_custody","Separate goods receipt, stock adjustment, counting and custody.","R04, R11","High"),
 ("exec","gl_post · gl_close · recon_prep · fin_report","Separate journal posting from period/year close from recon prep and read-only reporting.","R05, R06","High"),
 ("masterdata","md_vendor · md_item · md_config","Separate vendor master from item/config master.","R02, R13","Medium"),
 ("crm","(unchanged) — resolved by ROLE separation","Keep customer/credit master OFF any selling role rather than split the permission.","R09","Medium"),
 ("returns","(unchanged) — resolved by ROLE separation","Keep returns processing OFF any refund-issuing role.","R12","Medium"),
]
# 2) to-be roles (single-duty; SoD-clean). perms use new sub-permission tokens.
TB_ROLES=[
 ("Cashier",{"pos_sell"},"Sells only — cannot refund/void or reconcile the till."),
 ("POS Supervisor",{"pos_refund","pos_till"},"Authorizes refunds/voids and reconciles the till — does not ring sales."),
 ("AR Clerk",{"ar","order_mgt","claim_mgt","delivery"},"Posts AR / manages orders — no credit-master, no close, no approval."),
 ("AP Clerk",{"creditors"},"Pays vendors — cannot create vendors or raise POs."),
 ("Buyer (Procurement)",{"procurement"},"Raises PR/PO — cannot pay or receive goods."),
 ("Warehouse Operator",{"wh_receive","wh_custody","lots","locations","mobile","images"},"Receives & holds stock — cannot order, adjust, or count."),
 ("Inventory Controller",{"wh_adjust"},"Approves stock adjustments — independent of counting."),
 ("Stock Counter",{"wh_count"},"Performs physical counts — independent of adjustments."),
 ("GL Accountant",{"gl_post","recon_prep","fin_report"},"Posts JEs & prepares reconciliations — cannot close or certify."),
 ("Financial Controller",{"gl_close","approvals","fin_report"},"Closes periods, approves JEs & certifies recons — does not prepare/post."),
 ("Master Data Admin",{"md_vendor","md_item","md_config","bom_master"},"Maintains master data — no transactional rights."),
 ("Pricing Manager",{"pricelist","promos"},"Maintains prices/promotions — cannot sell."),
 ("CRM / Credit Manager",{"crm"},"Maintains customer & credit master — cannot enter orders."),
 ("Returns Clerk",{"returns"},"Processes returns — refund issuance authorized separately."),
 ("Access Administrator",{"users"},"Manages user access only — no transactional rights (isolated)."),
 ("Executive (read-only)",{"fin_report","dashboard","planner","marketing"},"Reporting & analytics only — no posting."),
 ("Customer",{"order_cust","cust_pos","cust_dash","loyalty","track"},"Portal self-service (unchanged)."),
 ("Superuser / Admin",{"__ALL__"},"Break-glass only — minimal named users, MFA, monitored (inherent superuser)."),
]
# 3) the 16 rules remapped to to-be sub-permissions
TB_RULES=[
 ("R01",{"users"},{"pos_sell","pos_refund","order_mgt","procurement","creditors","ar","returns","pricelist","promos","md_vendor","md_item","md_config","wh_receive","wh_adjust","gl_post"}),
 ("R02",{"md_vendor"},{"creditors"}),
 ("R03",{"procurement"},{"creditors"}),
 ("R04",{"procurement"},{"wh_receive"}),
 ("R05",{"gl_post"},{"gl_close"}),
 ("R06",{"recon_prep"},{"approvals"}),
 ("R07",{"procurement","pos_sell","ar","creditors","order_mgt"},{"approvals"}),
 ("R08",{"pos_sell"},{"pos_refund","pos_till"}),
 ("R09",{"crm","md_vendor"},{"pos_sell","order_mgt","order_cust"}),
 ("R10",{"pricelist","promos"},{"pos_sell","order_mgt"}),
 ("R11",{"wh_adjust"},{"wh_count"}),
 ("R12",{"returns"},{"pos_refund"}),
 ("R13",{"md_item","md_config","bom_master"},{"pos_sell","order_mgt","procurement","creditors","ar"}),
 ("R14",{"crm_reward"},{"pos_sell"}),
 ("R15",{"crm_points_adjust"},{"crm_member"}),
 ("R16",{"crm_campaign"},{"crm_points_adjust"}),
]
def tb_conflicts(perms):
    if "__ALL__" in perms: return [rid for rid,_,_ in TB_RULES]
    return [rid for rid,A,B in TB_RULES if (perms&A) and (perms&B)]

# ----- tab: Proposed Permissions -----
pp=wb.create_sheet("Proposed Permissions"); pp.sheet_view.showGridLines=False
pp["A1"]="Proposed permission splits (to-be) — single-duty sub-permissions"; pp["A1"].font=f(12,True,NAVY); pp.merge_cells("A1:E1")
hh=["Current permission","Proposed sub-permissions","Rationale","Resolves rules","Priority"]; wwp=[20,34,46,16,12]
for j,x in enumerate(hh,1):
    c=pp.cell(2,j,x); c.font=f(10,True,"FFFFFF"); c.fill=fill(STEEL); c.alignment=CTR; c.border=BORDER
    pp.column_dimensions[get_column_letter(j)].width=wwp[j-1]
for i,(old,new,why,res,pri) in enumerate(SPLITS,3):
    band=BAND if i%2==1 else "FFFFFF"
    for j,v in enumerate([old,new,why,res,pri],1):
        c=pp.cell(i,j,v); c.font=f(9); c.alignment=WRAP; c.border=BORDER; c.fill=fill(band)
        if j in (1,4,5): c.alignment=CTR
    fc,tc=SENS_STYLE[pri]; sc=pp.cell(i,5); sc.fill=fill(fc); sc.font=f(9,True,tc)
    pp.cell(i,1).font=f(9,True,NAVY)
pp.row_dimensions[2].height=28

# ----- tab: Proposed Roles -----
pr2=wb.create_sheet("Proposed Roles"); pr2.sheet_view.showGridLines=False
pr2["A1"]="Proposed role design (to-be) — single-duty roles"; pr2["A1"].font=f(12,True,NAVY); pr2.merge_cells("A1:C1")
hh=["Role","Permissions","SoD principle"]; wwr=[24,40,52]
for j,x in enumerate(hh,1):
    c=pr2.cell(2,j,x); c.font=f(10,True,"FFFFFF"); c.fill=fill(STEEL); c.alignment=CTR; c.border=BORDER
    pr2.column_dimensions[get_column_letter(j)].width=wwr[j-1]
for i,(name,perms,principle) in enumerate(TB_ROLES,3):
    band=BAND if i%2==1 else "FFFFFF"
    pset="ALL (break-glass)" if "__ALL__" in perms else ", ".join(sorted(perms))
    for j,v in enumerate([name,pset,principle],1):
        c=pr2.cell(i,j,v); c.font=f(9); c.alignment=WRAP; c.border=BORDER; c.fill=fill(band)
    pr2.cell(i,1).font=f(9,True,NAVY); pr2.cell(i,1).alignment=CTR
pr2.row_dimensions[2].height=24

# ----- tab: Remediated Matrix -----
rm=wb.create_sheet("Remediated Matrix"); rm.sheet_view.showGridLines=False
rm["A1"]="Remediated conflict matrix (to-be design) — target: zero conflicts except inherent superuser"; rm["A1"].font=f(11,True,NAVY)
rm.merge_cells("A1:P1")
rm.cell(2,1,"Role"); rm.cell(2,1).font=f(10,True,"FFFFFF"); rm.cell(2,1).fill=fill(NAVY); rm.cell(2,1).border=BORDER; rm.cell(2,1).alignment=CTR
rm.column_dimensions["A"].width=22
for j,(rid,_,_) in enumerate(TB_RULES,2):
    c=rm.cell(2,j,rid); c.font=f(9,True,"FFFFFF"); c.fill=fill(NAVY); c.border=BORDER; c.alignment=CTR
    rm.column_dimensions[get_column_letter(j)].width=6
ccol2=len(TB_RULES)+2
c=rm.cell(2,ccol2,"# Conflicts"); c.font=f(9,True,"FFFFFF"); c.fill=fill(NAVY); c.border=BORDER; c.alignment=CTR
rm.column_dimensions[get_column_letter(ccol2)].width=11
tb_total=0
for i,(name,perms,_) in enumerate(TB_ROLES,3):
    a=rm.cell(i,1,name); a.font=f(9,True,NAVY); a.border=BORDER; a.alignment=CTR
    hits=set(tb_conflicts(perms)); is_admin="__ALL__" in perms
    n=0
    for j,(rid,_,_) in enumerate(TB_RULES,2):
        cell=rm.cell(i,j);
        if is_admin: cell.value="▲"; cell.fill=fill(INH_F); cell.font=f(9,True,"FFFFFF")
        elif rid in hits: cell.value="✕"; cell.fill=fill(CONF_F); cell.font=f(9,True,"FFFFFF"); n+=1
        else: cell.value="·"; cell.fill=fill(CLEAR_F); cell.font=f(9,False,"006100")
        cell.border=BORDER; cell.alignment=CTR
    tot=rm.cell(i,ccol2); tot.border=BORDER; tot.alignment=CTR
    if is_admin: tot.value="ALL (▲)"; tot.font=f(9,True,HI_T)
    else: tot.value=n; tot.font=f(10,True,"006100" if n==0 else HI_T); tb_total+=n
note_r=len(TB_ROLES)+4
rm.cell(note_r,1,(f"Result: {tb_total} residual conflicts across the 17 operational roles (target 0). 'Superuser/Admin' remains an "
 "inherent superuser (break-glass) — mitigated by minimal named users, MFA, and the tamper-evident audit log + hash-chained "
 "journal. Compare with the 'Conflict Matrix' tab (current design: 18 conflicts)."))
rm.merge_cells(start_row=note_r,start_column=1,end_row=note_r+2,end_column=ccol2)
rm.cell(note_r,1).font=f(9,False,GREY); rm.cell(note_r,1).alignment=WRAP
rm.freeze_panes="B3"; rm.row_dimensions[2].height=16

# move Cover first
wb.move_sheet("Cover", -wb.sheetnames.index("Cover"))
out="compliance/Oshinei_ERP_SoD_Matrix_v1.xlsx"
wb.save(out)
print("TO-BE residual conflicts (non-admin):", sum(len(tb_conflicts(p)) for _,p,_ in TB_ROLES if "__ALL__" not in p))
# console summary
print("WROTE",out)
for role in ROLE_ORDER:
    rc=conflicts_for(ROLES[role])
    if role=="Admin":
        print(f"  {role:12s}: ALL (inherent superuser)")
    else:
        ids=",".join(x[0] for x in rc)
        print(f"  {role:12s}: {len(rc)} conflicts [{ids}]")
print("Total non-admin detected conflicts:", len(det_nonadmin))
