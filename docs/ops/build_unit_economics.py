"""
Generate docs/ops/unit-economics-model.xlsx — an ILLUSTRATIVE, formula-driven unit-economics model
(PwC Capital Markets follow-up). Inputs live on the 'Assumptions' sheet; the 'Model' sheet references them
with live Excel formulas, so overwriting an assumption recomputes every downstream metric.

This is a GENERATED binary — regenerate, don't hand-edit:  python3 docs/ops/build_unit_economics.py
Pairs with docs/ops/unit-economics-model.md (narrative) + docs/ops/pricing-and-ai-cogs.md (price list).
All figures are clearly-labeled illustrative defaults; replace with live data where a source is named.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Arial"
NAVY = "1F4E78"; STEEL = "2E5C8A"; BAND = "F2F6FB"; AMBER = "FFEB9C"; GREEN = "C6EFCE"; RED = "FFC7CE"

def f(sz=10, bold=False, color="000000"): return Font(name=FONT, size=sz, bold=bold, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
RIGHT = Alignment(horizontal="right")

wb = Workbook()

# ───────────────────────── Assumptions ─────────────────────────
A = wb.active; A.title = "Assumptions"
A.sheet_view.showGridLines = False
A.column_dimensions["A"].width = 38; A.column_dimensions["B"].width = 16; A.column_dimensions["C"].width = 62
A["A1"] = "Unit-economics assumptions (ILLUSTRATIVE — replace with real data)"
A["A1"].font = f(14, True, NAVY); A.merge_cells("A1:C1")
A["A2"] = "Overwrite column B; the Model sheet recomputes. THB unless noted."; A["A2"].font = f(9, False, STEEL); A.merge_cells("A2:C2")

# (label, value, fmt, note) — row index starts at 4
rows = [
    ("FX rate (THB per USD)", 35, "0", "set to current"),
    ("Anthropic blended token cost (USD / 1M tokens)", 4.00, "0.00", "Sonnet-reasoning + Haiku-relay mix w/ prompt caching; ai-models.ts tiering"),
    ("Token cost (THB / 1k tokens)", "=B5/1000*B4", "0.000", "derived = USD/1M ÷ 1000 × FX"),
    ("Infra base (THB / mo, platform)", 52500, "#,##0", "≈ $1,500/mo Alibaba Cloud Bangkok, lean"),
    ("Paying tenants", 50, "0", "replace with live count"),
    ("Infra allocation (THB / tenant / mo)", "=B7/B8", "#,##0", "derived = infra base ÷ tenants (falls as tenants grow)"),
    ("Starter price (THB / mo)", 990, "#,##0", "billing PLAN_SEED"),
    ("Pro price (THB / mo)", 2900, "#,##0", "billing PLAN_SEED"),
    ("Enterprise ARPU (THB / mo)", 15000, "#,##0", "custom-priced; illustrative"),
    ("Pro avg AI usage (tokens / day)", 60000, "#,##0", "30% of 200k included; replace w/ ai_token_usage actuals"),
    ("Enterprise avg AI usage (tokens / day)", 600000, "#,##0", "30% of 2M included; replace w/ ai_token_usage actuals"),
    ("Plan mix — Starter", 0.20, "0%", "share of paying tenants"),
    ("Plan mix — Pro", 0.70, "0%", "share of paying tenants"),
    ("Plan mix — Enterprise", 0.10, "0%", "share of paying tenants"),
    ("Blended CAC (THB)", 4500, "#,##0", "self-serve + light sales; illustrative"),
    ("Monthly logo churn", 0.03, "0.0%", "illustrative SMB; saas-metrics churn.churn_rate_30d_pct"),
    ("Target gross margin", 0.75, "0%", "guardrail"),
]
A["A3"] = "Input"; A["B3"] = "Value"; A["C3"] = "Source / note"
for c in ("A3", "B3", "C3"): A[c].font = f(10, True, "FFFFFF"); A[c].fill = fill(STEEL)
r0 = 4
for i, (lab, val, fmt, note) in enumerate(rows):
    r = r0 + i
    A[f"A{r}"] = lab; A[f"A{r}"].font = f(10); A[f"A{r}"].alignment = WRAP
    A[f"B{r}"] = val; A[f"B{r}"].font = f(10, True, NAVY); A[f"B{r}"].number_format = fmt; A[f"B{r}"].fill = fill(BAND); A[f"B{r}"].alignment = RIGHT
    A[f"C{r}"] = note; A[f"C{r}"].font = f(9, False, "606060"); A[f"C{r}"].alignment = WRAP
    for col in ("A", "B", "C"): A[f"{col}{r}"].border = BORDER

# named refs into Assumptions!B*
FX, TOK_USD, TOK_THB, INFRA, TEN, ALLOC, P_S, P_P, ARPU_E, U_P, U_E, M_S, M_P, M_E, CAC, CHURN, TGM = (f"Assumptions!$B${r0+i}" for i in range(len(rows)))

# ───────────────────────── Model ─────────────────────────
M = wb.create_sheet("Model")
M.sheet_view.showGridLines = False
for col, w in zip("ABCDEFG", (30, 14, 14, 14, 14, 14, 12)): M.column_dimensions[col].width = w
M["A1"] = "Unit economics — computed (live formulas)"; M["A1"].font = f(14, True, NAVY); M.merge_cells("A1:G1")

# Per-plan economics table
M["A3"] = "Per-plan economics"; M["A3"].font = f(11, True, NAVY)
hdr = ["Plan", "Price/mo", "AI COGS/mo", "Infra alloc", "Total COGS", "Gross profit", "Gross margin"]
for j, h in enumerate(hdr):
    c = M.cell(row=4, column=1 + j, value=h); c.font = f(10, True, "FFFFFF"); c.fill = fill(STEEL); c.alignment = WRAP; c.border = BORDER
# Starter (AI off), Pro, Enterprise
plan_rows = [
    ("Starter", P_S, "0", U_P),       # AI off → 0 COGS
    ("Pro", P_P, f"={U_P}*30/1000*{TOK_THB}", U_P),
    ("Enterprise", ARPU_E, f"={U_E}*30/1000*{TOK_THB}", U_E),
]
for i, (name, price, aicogs, _u) in enumerate(plan_rows):
    r = 5 + i
    M.cell(row=r, column=1, value=name).font = f(10, True)
    M.cell(row=r, column=2, value=f"={price}").number_format = "#,##0"
    M.cell(row=r, column=3, value=(0 if name == "Starter" else aicogs)).number_format = "#,##0"
    M.cell(row=r, column=4, value=f"={ALLOC}").number_format = "#,##0"
    M.cell(row=r, column=5, value=f"=C{r}+D{r}").number_format = "#,##0"
    M.cell(row=r, column=6, value=f"=B{r}-E{r}").number_format = "#,##0"
    gm = M.cell(row=r, column=7, value=f"=IF(B{r}=0,0,F{r}/B{r})"); gm.number_format = "0%"
    for col in range(1, 8):
        cell = M.cell(row=r, column=col); cell.border = BORDER
        if col == 7:
            cell.fill = fill(GREEN if name != "Starter" else AMBER)

# Blended SaaS metrics
M["A10"] = "Blended SaaS metrics"; M["A10"].font = f(11, True, NAVY)
metrics = [
    ("Blended ARPU (THB/mo)", f"={M_S}*{P_S}+{M_P}*{P_P}+{M_E}*{ARPU_E}", "#,##0", "Σ(mix×price); saas-metrics revenue.arpu"),
    ("MRR (THB)", f"=B11*{TEN}", "#,##0", "ARPU × paying tenants; saas-metrics revenue.mrr"),
    ("ARR (THB)", "=B12*12", "#,##0", "MRR × 12; saas-metrics revenue.arr"),
    ("Blended COGS / tenant (THB/mo)", f"={M_S}*({ALLOC})+{M_P}*(C6+{ALLOC})+{M_E}*(C7+{ALLOC})", "#,##0", "mix-weighted infra + AI COGS"),
    ("Blended gross margin", "=(B11-B14)/B11", "0%", "(ARPU − blended COGS) / ARPU"),
    ("CAC payback (months)", f"={CAC}/(B11*B15)", "0.0", "CAC / (ARPU × GM)"),
    ("LTV (THB)", f"=(B11*B15)/{CHURN}", "#,##0", "(ARPU × GM) / monthly churn"),
    ("LTV : CAC", f"=B17/{CAC}", "0.0", "LTV / CAC (≥3 healthy)"),
]
M["A11"]  # anchor
hr = 11
M.cell(row=hr - 0, column=1)  # noop
M.cell(row=10, column=1)
# header for metrics
M.cell(row=10, column=2)
for i, (lab, formula, fmt, note) in enumerate(metrics):
    r = 11 + i
    M.cell(row=r, column=1, value=lab).font = f(10)
    v = M.cell(row=r, column=2, value=formula); v.number_format = fmt; v.font = f(10, True, NAVY); v.fill = fill(BAND); v.alignment = RIGHT
    M.cell(row=r, column=3, value=note).font = f(9, False, "606060")
    M.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    for col in (1, 2): M.cell(row=r, column=col).border = BORDER

M["A20"] = ("Guardrail: no unlimited AI tier (finite ai_tokens_daily_max per plan); overage above the included "
            "cap is billed (ai_overage_rate_thb_per_1k); off-Opus tiering keeps blended token COGS ~0.14 THB/1k. "
            "Watch Starter — margin-negative on infra allocation until tenant count amortizes the base.")
M["A20"].font = f(9, False, "606060"); M.merge_cells("A20:G22"); M["A20"].alignment = WRAP

import os
out = os.path.join(os.path.dirname(__file__), "unit-economics-model.xlsx")
wb.save(out)
print(f"WROTE {out}")
